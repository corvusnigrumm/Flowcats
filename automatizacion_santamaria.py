# -*- coding: utf-8 -*-
"""
automatizacion_santamaria.py
====================
Extrae UN articulo reciente por categoria de https://www.eltiempo.com/ y 
https://www.portafolio.co/ y los exporta a libros Excel separados:
    - El Tiempo.xlsx
    - Portafolio.xlsx

Caracteristicas:
    - Titulo corto inteligente (max 2 palabras con sentido coherente)
    - Casilla de verificacion que resalta la fila en verde (solo admite ✔)
    - Diseño en fuente Century Gothic con encabezados de fondo negro.
    - Filtro de noticias exclusivas para suscriptores (solo El Tiempo)
    - Categorias agrupadas para Portafolio (5 grupos)
    - Columna de resumen corto al final

Dependencias:
    pip install requests openpyxl lxml

Uso:
    python automatizacion_santamaria.py
"""

import sys
import io
import re
import os
import json
from datetime import datetime
from email.utils import parsedate_to_datetime

from typing import Dict, TypedDict

import requests  # type: ignore
import openpyxl  # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
from openpyxl.formatting.rule import FormulaRule  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore

try:
    from lxml import etree as ET  # type: ignore
except ImportError:
    import xml.etree.ElementTree as ET

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

# ─────────────────────────────────────────────
# FEEDS RSS Y CONFIGURACION DE SITIOS
# ─────────────────────────────────────────────

class SiteConfig(TypedDict):
    output_file: str
    url_base: str
    feeds: Dict[str, str]
    # feeds_grouped es alternativo: Dict[str, list[str]] para categorias agrupadas

# Definición de los dos diarios a extraer
SITES: Dict[str, SiteConfig] = {
    "El Tiempo": {
        "output_file": "El Tiempo.xlsx",
        "url_base": "eltiempo.com",
        "feeds": {
            "Bogota":   "https://www.eltiempo.com/rss/bogota.xml",
            "Colombia": "https://www.eltiempo.com/rss/colombia.xml",
            "Mundo":    "https://www.eltiempo.com/rss/mundo.xml",
            "Economia": "https://www.eltiempo.com/rss/economia.xml",
            "Deportes": "https://www.eltiempo.com/rss/deportes.xml",
            "Politica": "https://www.eltiempo.com/rss/politica.xml",
            "Cultura":  "https://www.eltiempo.com/rss/cultura.xml",
            "Vida":     "https://www.eltiempo.com/rss/vida.xml",
            "Justicia": "https://www.eltiempo.com/rss/justicia.xml",
            "Salud":    "https://www.eltiempo.com/rss/salud.xml",
        }
    },
}

TREND_KEYWORDS = {
    "crisis", "reforma", "elecciones", "inflacion", "inflacion", "dolar", "dolar",
    "subsidio", "impuestos", "pension", "pension", "salario", "empleo", "desempleo",
    "salud", "seguridad", "justicia", "guerra", "arancel", "precio", "precios",
    "petroleo", "petroleo", "tecnologia", "tecnologia", "inteligencia artificial",
    "ia", "seleccion", "seleccion", "futbol", "futbol", "escandalo", "escandalo",
    "historico", "historico", "record", "record", "viral", "denuncia"
}

SEO_KEYWORDS = {
    "como", "como", "que", "que", "por que", "por que", "claves", "guia", "guia",
    "paso a paso", "explicamos", "esto significa", "fechas", "requisitos",
    "quienes", "quienes", "cuando", "cuando", "precio", "precios", "ranking"
}

CATEGORY_POTENTIAL = {
    "Colombia": 12,
    "Politica": 12,
    "Economia": 11,
    "Salud": 11,
    "Bogota": 9,
    "Mundo": 8,
    "Justicia": 8,
    "Vida": 7,
    "Deportes": 7,
    "Cultura": 5,
}

# Portafolio usa feeds agrupados: cada categoria muestra una sola etiqueta 
# pero recoge articulos de multiples feeds RSS.
PORTAFOLIO_CONFIG = {
    "output_file": "Portafolio.xlsx",
    "url_base": "portafolio.co",
    "feeds_grouped": {
        "Economía, indicadores": [
            "https://www.portafolio.co/rss/economia.xml",
            "https://www.portafolio.co/rss/indicadores-economicos.xml",
        ],
        "Tendencias": [
            "https://www.portafolio.co/rss/tendencias.xml",
        ],
        "Internacional, emprendimiento, tecno": [
            "https://www.portafolio.co/rss/internacional.xml",
            "https://www.portafolio.co/rss/negocios/emprendimiento.xml",
            "https://www.portafolio.co/rss/innovacion.xml",
        ],
        "Energía, sostenibilidad": [
            "https://www.portafolio.co/rss/energia.xml",
            "https://www.portafolio.co/rss/sostenibilidad.xml",
        ],
        "Negocios, mis finanzas": [
            "https://www.portafolio.co/rss/negocios.xml",
            "https://www.portafolio.co/rss/mis-finanzas.xml",
        ],
    }
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/rss+xml, application/xml, text/xml, */*",
    "Accept-Language": "es-CO,es;q=0.9",
}

GEMINI_API_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent"
GEMINI_TIMEOUT = 25

# Color de texto por categorias (se comparten algunas, otras usan general)
CATEGORY_COLORS = {
    # El Tiempo
    "Opinion":    "4E342E",
    "Politica":   "B71C1C",
    "Economia":   "0D47A1",
    "Deportes":   "1B5E20",
    "Mundo":      "4A148C",
    "Bogota":     "E65100",
    "Colombia":   "006064",
    "Cultura":    "00695C",
    "Vida":       "33691E",
    "Justicia":   "37474F",
    "Salud":      "880E4F",
    # Portafolio (categorias agrupadas)
    "Economía, indicadores":              "0D47A1",
    "Tendencias":                          "00838F",
    "Internacional, emprendimiento, tecno": "4A148C",
    "Energía, sostenibilidad":             "D84315",
    "Negocios, mis finanzas":              "1565C0",
    # General
    "General":    "424242",
}

PREFIX_NOISE = [
    r"^en vivo\s*[:\-\—\–]",
    r"^en directo\s*[:\-\—\–]",
    r"^video\s*[:\-\—\–]",
    r"^fotos\s*[:\-\—\–]",
    r"^en fotos\s*[:\-\—\–]",
    r"^galería\s*[:\-\—\–]",
    r"^atención\s*[:\-\—\–]",
    r"^urgente\s*[:\-\—\–]",
    r"^análisis\s*[:\-\—\–]",
    r"^opinión\s*[:\-\—\–]",
    r"^ojo\s*[:\-\—\–]",
    r"^minuto a minuto\s*[:\-\—\–]",
    r"^lo último\s*[:\-\—\–]",
]

HANGING_WORDS = {
    "de", "del", "en", "para", "con", "por", "y", "e", "o", "u", "a", "al",
    "tras", "el", "la", "los", "las", "un", "una", "unos", "unas", "su", "sus",
    "sin", "sobre", "entre", "hacia", "hasta", "desde", "ante", "bajo", "contra",
    "que", "se", "es", "como", "mas", "más", "pero", "sino", "este", "esta",
    "estos", "estas", "ese", "esa", "esos", "esas", "us", "usd", "durante",
    "según", "segun", "incluyendo", "solo", "sólo", "cada", "aunque", "porque",
    "cuando", "donde", "mientras", "tan", "muy"
}

STOP_WORDS = {
    "de", "del", "en", "para", "con", "por", "y", "e", "o", "u", "a", "al",
    "tras", "el", "la", "los", "las", "un", "una", "unos", "unas", "su", "sus",
    "sin", "sobre", "entre", "hacia", "hasta", "desde", "ante", "bajo", "contra",
    "que", "se", "es", "como", "mas", "más", "pero", "sino", "este", "esta",
    "estos", "estas", "ese", "esa", "esos", "esas", "durante", "según", "segun",
    "solo", "sólo", "cada", "aunque", "porque", "cuando", "donde", "mientras", "tan", "muy"
}

# ─────────────────────────────────────────────
# UTILIDADES Y PROCESAMIENTO
# ─────────────────────────────────────────────

def clean_and_format_title(raw_title: str, max_words: int = 5) -> str:
    """
    Formatea el titular periodístico para Flowcards cuando no hay IA:
    - Conserva las palabras clave y conectores indispensables para mantener la coherencia y sentido (máx max_words palabras).
    - Evita cortar palabras a la mitad o dejar frases incomprensibles.
    """
    if not raw_title or not str(raw_title).strip():
        return "Sin título"
    
    text = str(raw_title).strip()
    
    # 1. Quitar prefijos periodísticos comunes
    for pattern in PREFIX_NOISE:
        text = re.sub(pattern, "", text, flags=re.IGNORECASE).strip()
        
    # 2. Quitar citas/autorías al final
    text = re.sub(r':\s*[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+){1,3}\s*$', '', text)
    text = re.sub(r',\s*según\s+.*$', '', text, flags=re.IGNORECASE)
    
    # 3. Limpiar signos extraños
    text = re.sub(r'[^\w\sÁÉÍÓÚáéíóúñÑüÜ.,;:\-]', ' ', text)
    
    words = [w.strip() for w in text.split() if w.strip()]
    if not words:
        return "Sin título"
        
    selected = words[:max_words]
    # Quitar conectores que queden colgando al final si la frase se recorta
    while selected and selected[-1].lower() in HANGING_WORDS and len(selected) > 2:
        selected.pop()
        
    res = " ".join(selected)
    if res:
        res = res[0].upper() + res[1:]
    return res

def shorten_title_smart(title: str, max_words: int = 10) -> str:
    """Función de compatibilidad que llama a clean_and_format_title."""
    return clean_and_format_title(title, max_words=max_words)

def parse_pubdate(pubdate_str: str):
    if not pubdate_str:
        dt = datetime.now()
        local = dt.astimezone()
        hora_12 = local.strftime("%I:%M %p").lstrip("0")
        return local.strftime("%Y-%m-%d"), hora_12, local.timestamp()
    try:
        dt = parsedate_to_datetime(pubdate_str)
    except Exception:
        try:
            dt = datetime.fromisoformat(pubdate_str)
        except Exception:
            dt = datetime.now()
            local = dt.astimezone()
            hora_12 = local.strftime("%I:%M %p").lstrip("0")
            return local.strftime("%Y-%m-%d"), hora_12, local.timestamp()
    local = dt.astimezone()
    hora_12 = local.strftime("%I:%M %p").lstrip("0")
    return local.strftime("%Y-%m-%d"), hora_12, local.timestamp()

def make_amp_url(url: str, url_base: str) -> str:
    """Inserta /amp/ en la URL específica del diario."""
    if not url:
        return url
    target = url_base + "/"
    amp_target = url_base + "/amp/"
    return url.replace(target, amp_target, 1)


def analyze_seo_potential(article: dict) -> dict:
    """Asigna un puntaje heuristico de potencial SEO/tendencia para El Tiempo."""
    titulo_raw = str(article.get("titulo_raw", "") or "")
    resumen = str(article.get("resumen", "") or "")
    categoria = str(article.get("categoria", "") or "")
    combined_text = f"{titulo_raw} {resumen}".lower()

    score = 35
    reasons = []

    age_hours = max(0.0, (datetime.now().astimezone().timestamp() - float(article.get("timestamp", 0))) / 3600)
    if age_hours <= 6:
        score += 16
        reasons.append("muy reciente")
    elif age_hours <= 12:
        score += 12
        reasons.append("reciente")
    elif age_hours <= 24:
        score += 8
        reasons.append("aun fresca")
    elif age_hours <= 48:
        score += 4

    category_bonus = CATEGORY_POTENTIAL.get(categoria, 4)
    score += category_bonus
    if category_bonus >= 10:
        reasons.append(f"tema masivo de {categoria.lower()}")
    elif category_bonus >= 7:
        reasons.append(f"categoria activa: {categoria.lower()}")

    trend_hits = sorted({kw for kw in TREND_KEYWORDS if kw in combined_text})
    if trend_hits:
        score += min(18, len(trend_hits) * 5)
        reasons.append("tema con traccion de actualidad")

    seo_hits = sorted({kw for kw in SEO_KEYWORDS if kw in combined_text})
    if seo_hits:
        score += min(14, len(seo_hits) * 4)
        reasons.append("angulo util para busqueda")

    if re.search(r"\b\d+\b", titulo_raw):
        score += 7
        reasons.append("titulo con dato concreto")

    title_words = len([w for w in re.split(r"\s+", titulo_raw.strip()) if w])
    if 7 <= title_words <= 15:
        score += 7
        reasons.append("titulo con longitud competitiva")
    elif 5 <= title_words <= 18:
        score += 4

    title_chars = len(titulo_raw)
    if 45 <= title_chars <= 110:
        score += 6
    elif 30 <= title_chars <= 130:
        score += 3

    summary_len = len(resumen)
    if 90 <= summary_len <= 260:
        score += 5
    elif summary_len >= 45:
        score += 2

    if categoria == "Opinion":
        score -= 8

    score = max(0, min(100, score))

    if score >= 78:
        level = "Muy alto"
    elif score >= 63:
        level = "Alto"
    elif score >= 48:
        level = "Medio"
    else:
        level = "Bajo"

    seo_reason = ", ".join(reasons[:3]) if reasons else "senal limitada para tendencia o SEO"
    return {
        "seo_score": score,
        "seo_level": level,
        "seo_reason": seo_reason,
    }


def load_dotenv_simple():
    """Carga variables desde el archivo .env si existe."""
    env_paths = [".env", os.path.join(os.path.dirname(__file__), ".env")]
    for p in env_paths:
        if os.path.exists(p):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith("#") and "=" in line:
                            k, v = line.split("=", 1)
                            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
            except Exception:
                pass

load_dotenv_simple()

GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"
GROQ_MODEL = "openai/gpt-oss-120b"
GROQ_TIMEOUT = 25

def get_groq_api_key() -> str:
    load_dotenv_simple()
    return (
        os.environ.get("GROQ_API_KEY")
        or os.environ.get("GROQ_KEY")
        or ""
    ).strip()

def analyze_with_groq(article: dict) -> dict | None:
    api_key = get_groq_api_key()
    if not api_key:
        return None

    raw_title = article.get("titulo_raw", "")
    category = article.get("categoria", "")
    summary = article.get("resumen", "")

    # Bug 2 fix: prompt reescrito para titular con sentido real, no solo keywords
    prompt = (
        f"Eres un editor periodístico senior de El Tiempo / Portafolio.\n"
        f"Tu tarea es generar un TITULAR PERIODÍSTICO para Flowcard de MÁXIMO 5 PALABRAS.\n\n"
        f"DATOS DE LA NOTICIA:\n"
        f"- Categoría: {category}\n"
        f"- Titular original: {raw_title}\n"
        f"- Resumen: {summary}\n\n"
        f"REGLAS OBLIGATORIAS PARA EL TITULAR FLOWCARD:\n"
        f"1. MÁXIMO 5 PALABRAS en total (incluyendo conectores si son necesarios para el sentido).\n"
        f"2. El titular debe tener SENTIDO COMPLETO: incluye las keywords más importantes de la noticia.\n"
        f"3. Debe ser un titular PERIODÍSTICO de alto impacto, no solo una lista de palabras.\n"
        f"4. Ejemplos excelentes: 'Dólar baja a mínimos', 'Reforma pensional en la Corte',\n"
        f"   'Bogotá acelera visas turistas', 'Ballenas en Buenaventura', 'Petro anuncia cambios'.\n"
        f"5. Ejemplos MALOS (evitar): 'Dólar Mínimos' (sin sentido), 'Reforma Pensional Corte' (telegráfico).\n"
        f"6. Evalúa el potencial SEO y tendencia (0-100).\n\n"
        f'Responde ÚNICAMENTE en formato JSON estructurado:\n'
        f'{{\n'
        f'  "titulo_flowcard": "Titular con sentido",\n'
        f'  "seo_score": 88,\n'
        f'  "seo_level": "Muy alto",\n'
        f'  "seo_reason": "Motivo del impacto periodístico",\n'
        f'  "keyword_objetivo": "término clave",\n'
        f'  "trend_type": "Nacional / Economía / Tendencia"\n'
        f'}}'
    )

    try:
        from groq import Groq
        client = Groq(api_key=api_key)
        
        models_to_try = [
            "openai/gpt-oss-120b",
            "openai/gpt-oss-20b",
            "qwen/qwen3.8-27b",
            "qwen/qwen3.6-27b"
        ]
        
        for m in models_to_try:
            try:
                chat_completion = client.chat.completions.create(
                    model=m,
                    messages=[
                        {"role": "system", "content": "Eres un editor periodístico experto en titulares de alto impacto. Generas titulares breves (máx 5 palabras) con sentido completo y respondes únicamente en formato JSON."},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.3,
                    max_completion_tokens=512,
                    response_format={"type": "json_object"}
                )
                content_resp = chat_completion.choices[0].message.content
                if content_resp:
                    data_json = json.loads(content_resp)
                    score = int(data_json.get("seo_score", 80))
                    score = max(0, min(100, score))
                    raw_tf = str(data_json.get("titulo_flowcard", "")).strip()

                    # Bug 1 fix: NO aplicar clean_and_format_title() sobre el resultado de la IA.
                    # Solo limpiar caracteres extraños y aplicar title case sin recortar palabras.
                    titulo_flow = re.sub(r'[^\w\sÁÉÍÓÚáéíóúñÑüÜ.,;:\-]', '', raw_tf).strip()
                    titulo_flow = " ".join(titulo_flow.split())  # normalizar espacios
                    if titulo_flow:
                        titulo_flow = titulo_flow[0].upper() + titulo_flow[1:]  # capitalizar primera letra

                    # Fallback si la IA devuelve algo vacío o sin sentido (<2 palabras)
                    if not titulo_flow or len(titulo_flow.split()) < 2:
                        titulo_flow = clean_and_format_title(raw_title, max_words=5)
                    
                    return {
                        "seo_score": score,
                        "seo_level": str(data_json.get("seo_level", "Alto")).strip(),
                        "seo_reason": str(data_json.get("seo_reason", ""))[:140] or "Análisis SEO con IA",
                        "keyword_objetivo": str(data_json.get("keyword_objetivo", ""))[:60],
                        "trend_type": str(data_json.get("trend_type", ""))[:40],
                        "titulo_flowcard": titulo_flow,
                        "analysis_source": f"Groq ({m})",
                    }
            except Exception:
                continue
    except Exception:
        pass

    return None

def shorten_to_25_chars(title: str) -> str:
    """Asegura de forma estricta que el titular tenga MÁXIMO 25 caracteres y no termine en palabras colgantes o puntuación suelta."""
    if not title:
        return "Sin título"
    clean = re.sub(r'[^\w\sÁÉÍÓÚáéíóúñÑüÜ.,;:\-]', '', str(title).strip()).strip()
    clean = clean.rstrip(".,;: -")
    if len(clean) <= 25:
        words = clean.split()
        while words and words[-1].lower() in HANGING_WORDS and len(words) > 1:
            words.pop()
        return " ".join(words) if words else clean

    words = clean.split()
    result = []
    current_len = 0
    for w in words:
        projected = current_len + len(w) + (1 if result else 0)
        if projected > 25:
            break
        result.append(w)
        current_len = projected
    # Eliminar conectores que hayan quedado truncados al final
    while result and result[-1].lower() in HANGING_WORDS and len(result) > 1:
        result.pop()
    if not result:
        return clean[:25].rstrip(" .,;:-")
    return " ".join(result).rstrip(".,;:-")

def generate_temas_del_dia_headline(article: dict) -> str:
    """Genera un titular para 'Temas del Día' de MÁXIMO 25 caracteres totales con IA Groq.
    Bug 3 fix: prompt mejorado para garantizar coherencia periodística (verbo+sujeto, keywords, sin cortes).
    """
    api_key = get_groq_api_key()
    raw_title = article.get("titulo_raw", "")
    summary = article.get("resumen", "")
    category = article.get("categoria", "")
    
    if api_key:
        prompt = (
            f"Eres un editor de portadas periodísticas de El Tiempo / Portafolio.\n"
            f"Crea un TITULAR BREVE DE PORTADA para 'Temas del Día'.\n\n"
            f"REGLAS CRÍTICAS INVIOLABLES:\n"
            f"1. LONGITUD: ENTRE 15 Y 24 CARACTERES TOTALES (contando espacios y letras). NUNCA superar 25 caracteres.\n"
            f"2. SENTIDO COMPLETO: Debe tener sujeto + verbo o ser una frase periodística coherente y natural.\n"
            f"3. KEYWORDS: Incluye la palabra clave principal de la noticia (economía, política, seguridad, etc.).\n"
            f"4. NUNCA termines en preposiciones como 'de', 'en', 'a', 'por', 'con', 'para'.\n"
            f"5. EJEMPLOS PERFECTOS (fíjate en el conteo de caracteres):\n"
            f"   - 'Dólar baja a mínimos' (20 caracteres)\n"
            f"   - 'Petro anuncia reformas' (22 caracteres)\n"
            f"   - 'Alerta roja en Bogotá' (22 caracteres)\n"
            f"   - 'Tasas de interés caen' (22 caracteres)\n"
            f"   - 'Crisis política crece' (22 caracteres)\n"
            f"   - 'Corte frena reforma' (19 caracteres)\n\n"
            f"NOTICIA:\n"
            f"- Categoría: {category}\n"
            f"- Título original: {raw_title}\n"
            f"- Resumen: {summary[:250]}\n\n"
            f'Responde ÚNICAMENTE en formato JSON: {{"titular_25": "Titular Breve"}}'
        )
        try:
            from groq import Groq
            client = Groq(api_key=api_key)
            for m in ["openai/gpt-oss-120b", "openai/gpt-oss-20b", "qwen/qwen3.8-27b", "qwen/qwen3.6-27b"]:
                try:
                    resp = client.chat.completions.create(
                        model=m,
                        messages=[
                            {"role": "system", "content": "Eres un editor de portadas periodísticas. Creas titulares concisos de entre 15 y 24 caracteres con sentido completo y alto impacto."},
                            {"role": "user", "content": prompt}
                        ],
                        temperature=0.2,
                        max_completion_tokens=128,
                        response_format={"type": "json_object"}
                    )
                    content = resp.choices[0].message.content
                    if content:
                        parsed = json.loads(content)
                        t25 = str(parsed.get("titular_25", "")).strip()
                        t25 = t25.rstrip(".,;: -\"'")
                        # Validación: entre 8 y 25 caracteres y al menos 2 palabras
                        if t25 and len(t25.split()) >= 2:
                            return shorten_to_25_chars(t25)
                except Exception:
                    continue
        except Exception:
            pass

    # Fallback heurístico si no hay IA
    return shorten_to_25_chars(raw_title)

def export_temas_del_dia(articles: list, work_dir: str) -> str | None:
    """
    Selecciona las mejores 8 noticias de El Tiempo / Portafolio, genera para cada una un titular de <=25 caracteres con IA Groq
    y actualiza la plantilla 'TEMAS DEL DÍA.xlsx'.
    """
    if not articles:
        print("\n[!] No hay artículos para generar Temas del Día.")
        return None

    template_file = os.path.join(work_dir, "TEMAS DEL DÍA.xlsx")
    if not os.path.exists(template_file):
        template_file = os.path.join(work_dir, "TEMAS DEL DIA.xlsx")
        if not os.path.exists(template_file):
            template_file = "TEMAS DEL DÍA.xlsx"

    # Ordenar noticias por puntaje SEO o frescura y elegir las 8 mejores
    top_8 = sorted(
        articles,
        key=lambda x: (x.get("seo_score", 0), x.get("timestamp", 0)),
        reverse=True
    )[:8]

    print(f"\n--- PROCESANDO TEMAS DEL DÍA (Top 8 noticias con IA Groq) ---")

    try:
        if os.path.exists(template_file):
            wb = openpyxl.load_workbook(template_file)
            ws = wb.active
        else:
            # Crear libro con formato si no existe
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "TEMAS DEL DÍA"
            template_file = os.path.join(work_dir, "TEMAS DEL DÍA.xlsx")
            
            # Encabezados
            encabezados = ["RANKING", "CARACTERES", "TÍTULO (<= 25 chars)", "URL"]
            hdr_fill = PatternFill("solid", fgColor="000000")
            hdr_font = Font(name="Century Gothic", bold=True, color="FFFFFF", size=11)
            hdr_align = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25
            for col, texto in enumerate(encabezados, 1):
                cell = ws.cell(row=1, column=col, value=texto)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = hdr_align
                cell.border = thin_border()
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 35
            ws.column_dimensions["D"].width = 65

        for idx, art in enumerate(top_8, 1):
            row = idx + 2  # Filas 3 a 10
            titular_25 = generate_temas_del_dia_headline(art)
            art["titulo_temas_25"] = titular_25
            url_target = art.get("url_amp") or art.get("url_original", "")

            ws.cell(row=row, column=1, value=idx)                # Col A: RANKING
            ws.cell(row=row, column=2, value=f"=+LEN(C{row})")  # Col B: CARACTERES
            ws.cell(row=row, column=3, value=titular_25)         # Col C: TÍTULO (<= 25 chars)
            
            cell_url = ws.cell(row=row, column=4, value=url_target) # Col D: URL
            if url_target:
                cell_url.hyperlink = url_target
                cell_url.font = Font(name="Century Gothic", color="0277BD", underline="single", size=10)

            print(f"   Top {idx}: [{len(titular_25)} chars] \"{titular_25}\" -> {url_target[:50]}...")

        # Si hay menos de 8 noticias, limpiar filas sobrantes
        for idx in range(len(top_8) + 1, 9):
            row = idx + 2
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=f"=+LEN(C{row})")
            ws.cell(row=row, column=3, value="")
            ws.cell(row=row, column=4, value="")

        wb.save(template_file)
        print(f"\n[LISTO] Guardado 'TEMAS DEL DÍA.xlsx' con el Top 8 de noticias.")
        return template_file
    except PermissionError:
        print("\n[ERROR CRÍTICO] 'TEMAS DEL DÍA.xlsx' está abierto en Excel. Ciérrelo para actualizar.")
        return None
    except Exception as e:
        print(f"\n[ERROR] Falló la exportación de Temas del Día: {e}")
        return None


def is_subscriber_only(article_url: str) -> bool:
    """Verifica si un articulo de El Tiempo es exclusivo para suscriptores.
    Aplica filtros precisos para no confundir con menús JS globales."""
    try:
        resp = requests.get(article_url, headers=HEADERS, timeout=10)
        resp.raise_for_status()
        html_content = resp.text
        lower_html = html_content.lower()
        
        # 1. Coincidencias exactas en todo el HTML (para meta tags y variables JS)
        EXACT_MATCHES = [
            'content="csuscriptor-modal"',
            'content="exclusivo suscriptores"',
            '"premium": true', '"premium":true',
            'isrestricted":true'
        ]
        for match in EXACT_MATCHES:
            if match in lower_html:
                return True
                
        # 2. Análisis estructural detallado (solo dentro del contenido del artículo)
        if BeautifulSoup is not None:
            soup = BeautifulSoup(html_content, "html.parser")
            article_tag = soup.find("article")
            
            if article_tag:
                # Buscar un badge explícito de premium, pero EVITANDO las recomendaciones 
                # (elementos asides o cajas en el footer que traen c-suscriptor)
                header = article_tag.find("header") or article_tag
                
                # a. Clases CSS típicas del badge de suscriptor en el encabezado
                SUBSCRIBER_CLASSES = [
                    "badge-premium", "badge-subscriber", "premium-badge", "paywall"
                ]
                for css_class in SUBSCRIBER_CLASSES:
                    if header.find(class_=lambda c: c and css_class in c.lower()):
                        return True
                        
                # b. Atributos data- de paywall en el html local
                PAYWALL_ATTRS = ["data-premium", "data-paywall"]
                for attr in PAYWALL_ATTRS:
                    if header.find(attrs={attr: True}):
                        return True
                        
                # c. Texto ET visible en un badge pequeño (exclusivo ET) en el header
                for tag in header.find_all(["span", "div", "label"]):
                    # Verificar si la clase del tag sugiere un badge
                    tag_class = str(tag.get("class", "")).lower()
                    if tag.get_text(strip=True) == "ET" and ("badge" in tag_class or "label" in tag_class):
                        return True

    except Exception:
        pass  # Si falla, asumimos que NO es exclusivo
    return False


# ─────────────────────────────────────────────
# SCRAPING
# ─────────────────────────────────────────────

def fetch_top_article(feed_url: str, label: str, url_base: str, seen_urls: set, check_paywall: bool = False) -> dict | None:
    """Extrae el primer articulo valido de un feed RSS (el mas reciente).
    Si check_paywall=True, verifica que no sea exclusivo para suscriptores."""
    print(f"\n>> [{label}] Buscando en {feed_url}")
    try:
        resp = requests.get(feed_url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"   [!] Omitido (No accesible o no existe el feed).")
        return None

    try:
        root = ET.fromstring(resp.content)
    except ET.ParseError as exc:
        print(f"   [!] XML invalido.")
        return None

    items = root.findall(".//item")
    if not items:
        print(f"   [!] Feed vacío.")
        return None

    def _get_text(element) -> str:
        if element is not None and isinstance(element.text, str):
            val: str = element.text
            return val.strip()
        return ""

    parsed_items = []
    
    for item in items:
        link_el = item.find("link")
        art_url = _get_text(link_el)
        if not art_url:
            guid_el = item.find("guid")
            art_url = _get_text(guid_el)
            
        if not art_url:
            continue
            
        title_el = item.find("title")
        titulo_raw = _get_text(title_el)

        pub_el = item.find("pubDate")
        pub_str = _get_text(pub_el)
        fecha, hora, timestamp = parse_pubdate(pub_str)

        desc_el = item.find("description")
        resumen = _get_text(desc_el)

        cat_el = item.find("category")
        cat_raw = _get_text(cat_el)
        
        parsed_items.append({
            "art_url": art_url,
            "titulo_raw": titulo_raw,
            "fecha": fecha,
            "hora": hora,
            "timestamp": timestamp,
            "resumen": resumen,
            "cat_raw": cat_raw
        })

    # Ordenar chronológicamente descendente para asegurar que el primero es el más reciente
    parsed_items.sort(key=lambda x: x["timestamp"], reverse=True)

    for item_data in parsed_items:
        art_url = item_data["art_url"]

        if art_url in seen_urls:
            continue

        # Bug 6 fix: filtrar opinión, videos y podcasts en AMBOS medios
        cat_check = str(item_data.get("cat_raw", "") or "").lower()
        url_check = str(art_url or "").lower()
        titulo_check = str(item_data.get("titulo_raw", "") or "").lower()

        # Filtrar contenido de opinión
        if "opinion" in cat_check or "opinión" in cat_check or "/opinion/" in url_check:
            seen_urls.add(art_url)
            print(f"   [!] Omitido (Opinión): {item_data['titulo_raw'][:60]}")
            continue

        # Filtrar videos y podcasts
        if "/video/" in url_check or "/videos/" in url_check:
            seen_urls.add(art_url)
            print(f"   [!] Omitido (Video): {item_data['titulo_raw'][:60]}")
            continue
        if "/podcast/" in url_check or "/podcasts/" in url_check:
            seen_urls.add(art_url)
            print(f"   [!] Omitido (Podcast): {item_data['titulo_raw'][:60]}")
            continue

        # Filtrar por prefijo en el título (videos, fotos, galería, etc.) — ya cubiertos en PREFIX_NOISE
        if re.match(r'^(video|fotos|galería|podcast|en vivo|en directo)\s*[:\-]', titulo_check):
            seen_urls.add(art_url)
            print(f"   [!] Omitido (formato no válido): {item_data['titulo_raw'][:60]}")
            continue

        if check_paywall and is_subscriber_only(art_url):
            file_name_part = str(art_url.split('/')[-1])
            print(f"   [!] Omitido (Solo para suscriptores): {file_name_part[:60]}")
            seen_urls.add(art_url)
            continue

        seen_urls.add(art_url)
        
        # Mapeo por defecto o nombre del label
        categoria = label
        if label == "Opinion":
            cat_map = {"Columnistas": "Opinion", "Editorial": "Opinion", "Cartas": "Opinion", "Caricaturas": "Opinion"}
            categoria = cat_map.get(item_data["cat_raw"], label)
        titulo_3 = clean_and_format_title(item_data["titulo_raw"])
        url_amp  = make_amp_url(art_url, url_base)

        medio_name = "El Tiempo" if "eltiempo" in url_base.lower() else "Portafolio"
        article_data = {
            "titulo_3":     titulo_3,
            "titulo_raw":   item_data["titulo_raw"],
            "url_amp":      url_amp,
            "url_original": art_url,
            "categoria":    categoria,
            "fecha":        item_data["fecha"],
            "hora":         item_data["hora"],
            "timestamp":    item_data["timestamp"],
            "resumen":      item_data["resumen"],
            "medio":        medio_name,
            "source_name":  medio_name,
        }
        # Bug 4 fix: habilitar analyze_with_groq para AMBOS medios (El Tiempo y Portafolio)
        ai_analysis = analyze_with_groq(article_data)
        if ai_analysis:
            if ai_analysis.get("titulo_flowcard"):
                article_data["titulo_3"] = ai_analysis["titulo_flowcard"]
                titulo_3 = article_data["titulo_3"]
            article_data.update(ai_analysis)
        else:
            heuristic_analysis = analyze_seo_potential(article_data)
            heuristic_analysis["analysis_source"] = "Heuristico"
            article_data.update(heuristic_analysis)

        print(f"   OK -> \"{titulo_3}\" | {item_data['fecha']} {item_data['hora']}")
        return article_data

    print(f"   [!] Omitido (Todos los artículos de esta categoría ya fueron agregados o descartados).")
    return None


def fetch_top_from_group(feed_urls: list, label: str, url_base: str, seen_urls: set, check_paywall: bool = False) -> dict | None:
    """Busca en multiples feeds RSS y devuelve el articulo mas reciente no repetido.
    Usado para las categorias agrupadas de Portafolio.
    
    NOTA: Para evitar que los candidatos descartados bloqueen futuros grupos,
    se usa un conjunto temporal y solo el ganador se agrega al seen_urls global."""
    candidates = []
    candidate_urls = []  # URLs de los candidatos para revertir los no-ganadores
    
    for feed_url in feed_urls:
        # Usamos seen_urls normal: ya cubre URLs ya confirmadas en grupos anteriores
        art = fetch_top_article(feed_url, label, url_base, seen_urls, check_paywall=check_paywall)
        if art:
            candidates.append(art)
            candidate_urls.append(art["url_original"])
    
    if not candidates:
        return None
    
    # Ordenar por fecha+hora descendente y devolver el mas reciente
    candidates.sort(key=lambda x: x.get("timestamp", 0), reverse=True)
    winner = candidates[0]
    
    # Revertir del seen_urls los candidatos que NO ganaron,
    # para que otros grupos puedan considerar esos artículos si los necesitan.
    for art in candidates[1:]:
        seen_urls.discard(art["url_original"])
    
    return winner


# ─────────────────────────────────────────────
# EXCEL - DISEÑO Y EXPORTACIÓN
# ─────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color="B0BEC5")
    return Border(left=s, right=s, top=s, bottom=s)

def export_to_excel(articles: list, output_path: str, book_name: str, include_amp: bool = True) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = book_name
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 25

    is_el_tiempo = book_name == "El Tiempo"
    if include_amp:
        encabezados = ["✔", "Titulo", "URL SIN AMP", "CATEGORÍA", "Fecha", "URL AMP", "Resumen"]
    else:
        encabezados = ["✔", "Titulo", "URL SIN AMP", "CATEGORÍA", "Fecha", "Resumen"]
    if is_el_tiempo:
        encabezados.extend(["POTENCIAL", "PUNTAJE SEO", "MOTIVO SEO", "KEYWORD", "FUENTE ANALISIS"])

    hdr_fill = PatternFill("solid", fgColor="000000")
    hdr_font = Font(name="Century Gothic", bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")

    for col, texto in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col, value=texto)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border()

    last_column = get_column_letter(len(encabezados))
    ws.auto_filter.ref = f"A1:{last_column}1"

    fill_a = PatternFill("solid", fgColor="EEF2FA")
    fill_b = PatternFill("solid", fgColor="FFFFFF")

    if is_el_tiempo:
        sorted_arts = sorted(
            articles,
            key=lambda x: (x.get("seo_score", 0), x.get("timestamp", 0)),
            reverse=True,
        )
    else:
        sorted_arts = sorted(articles, key=lambda x: x.get("timestamp", 0), reverse=True)
    last_row = len(sorted_arts) + 1

    for r, art in enumerate(sorted_arts, 2):
        relleno = fill_a if r % 2 == 0 else fill_b
        fecha_hora = f"{art['fecha']}  {art['hora']}"

        # col_offset: cuando no hay AMP, las columnas a partir de Resumen se desplazan -1
        amp_shift = 0 if include_amp else -1

        ca = ws.cell(row=r, column=1, value="")
        ca.border = thin_border()
        ca.fill = relleno
        ca.font = Font(name="Century Gothic", bold=True, size=14, color="1B5E20")
        ca.alignment = Alignment(horizontal="center", vertical="center")

        cb = ws.cell(row=r, column=2, value=art["titulo_3"])
        cb.border = thin_border()
        cb.fill = relleno
        cb.font = Font(name="Century Gothic", bold=True, size=11)
        cb.alignment = Alignment(horizontal="left", vertical="center")

        cc = ws.cell(row=r, column=3, value=art["url_original"])
        cc.border = thin_border()
        cc.fill = relleno
        cc.alignment = Alignment(horizontal="left", vertical="center")
        if art["url_original"]:
            cc.hyperlink = art["url_original"]
            cc.font = Font(name="Century Gothic", color="0277BD", underline="single", size=10)
        else:
            cc.font = Font(name="Century Gothic", size=10)

        cat_color = CATEGORY_COLORS.get(art["categoria"], "424242")
        cd = ws.cell(row=r, column=4, value=art["categoria"])
        cd.border = thin_border()
        cd.fill = relleno
        cd.font = Font(name="Century Gothic", bold=True, color=cat_color, size=11)
        cd.alignment = Alignment(horizontal="center", vertical="center")

        ce = ws.cell(row=r, column=5, value=fecha_hora)
        ce.border = thin_border()
        ce.fill = relleno
        ce.font = Font(name="Century Gothic", size=10)
        ce.alignment = Alignment(horizontal="center", vertical="center")

        # Columna URL AMP (solo si include_amp=True)
        if include_amp:
            cf = ws.cell(row=r, column=6, value=art["url_amp"])
            cf.border = thin_border()
            cf.fill = relleno
            cf.alignment = Alignment(horizontal="left", vertical="center")
            if art["url_amp"]:
                cf.hyperlink = art["url_amp"]
                cf.font = Font(name="Century Gothic", color="1565C0", underline="single", size=10)
            else:
                cf.font = Font(name="Century Gothic", size=10)

        # Columna Resumen: col 7 con AMP, col 6 sin AMP
        cg = ws.cell(row=r, column=7 + amp_shift, value=art.get("resumen", ""))
        cg.border = thin_border()
        cg.fill = relleno
        cg.font = Font(name="Century Gothic", size=9, color="424242")
        cg.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if is_el_tiempo:
            seo_level = art.get("seo_level", "N/D")
            seo_score = art.get("seo_score", 0)
            seo_reason = art.get("seo_reason", "")

            level_color = "2E7D32"
            if seo_level == "Alto":
                level_color = "1565C0"
            elif seo_level == "Medio":
                level_color = "EF6C00"
            elif seo_level == "Bajo":
                level_color = "6D4C41"

            ch = ws.cell(row=r, column=8 + amp_shift, value=seo_level)
            ch.border = thin_border()
            ch.fill = relleno
            ch.font = Font(name="Century Gothic", bold=True, size=10, color=level_color)
            ch.alignment = Alignment(horizontal="center", vertical="center")

            ci = ws.cell(row=r, column=9 + amp_shift, value=seo_score)
            ci.border = thin_border()
            ci.fill = relleno
            ci.font = Font(name="Century Gothic", bold=True, size=10)
            ci.alignment = Alignment(horizontal="center", vertical="center")

            cj = ws.cell(row=r, column=10 + amp_shift, value=seo_reason)
            cj.border = thin_border()
            cj.fill = relleno
            cj.font = Font(name="Century Gothic", size=9, color="424242")
            cj.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            ck = ws.cell(row=r, column=11 + amp_shift, value=art.get("keyword_objetivo", ""))
            ck.border = thin_border()
            ck.fill = relleno
            ck.font = Font(name="Century Gothic", size=9, color="424242")
            ck.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            cl = ws.cell(row=r, column=12 + amp_shift, value=art.get("analysis_source", "Heuristico"))
            cl.border = thin_border()
            cl.fill = relleno
            cl.font = Font(name="Century Gothic", bold=True, size=9, color="424242")
            cl.alignment = Alignment(horizontal="center", vertical="center")

    dv = DataValidation(
        type="list",
        formula1='" ,✔"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.prompt = "Escribe ✔ para marcar"
    dv.promptTitle = "Marcar artículo"
    dv.error = "Solo se permite el signo: ✔ o celda vacía"
    dv.errorTitle = "Valor no válido"

    if last_row >= 2:
        dv.add(f"A2:A{last_row}")
        ws.add_data_validation(dv)

        green_fill = PatternFill("solid", fgColor="C8E6C9")
        green_font_bold = Font(name="Century Gothic", bold=True, color="1B5E20", size=11)
        rule = FormulaRule(
            formula=['LEN(TRIM($A2))>0'],
            fill=green_fill,
            font=green_font_bold,
        )
        ws.conditional_formatting.add(f"A2:{last_column}{last_row}", rule)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 22
    if include_amp:
        ws.column_dimensions["F"].width = 65   # URL AMP
        ws.column_dimensions["G"].width = 55   # Resumen
        if is_el_tiempo:
            ws.column_dimensions["H"].width = 16
            ws.column_dimensions["I"].width = 14
            ws.column_dimensions["J"].width = 42
            ws.column_dimensions["K"].width = 24
            ws.column_dimensions["L"].width = 18
    else:
        ws.column_dimensions["F"].width = 55   # Resumen (desplazado)
        if is_el_tiempo:
            ws.column_dimensions["G"].width = 16
            ws.column_dimensions["H"].width = 14
            ws.column_dimensions["I"].width = 42
            ws.column_dimensions["J"].width = 24
            ws.column_dimensions["K"].width = 18

    try:
        wb.save(output_path)
        print(f"\n[LISTO] Guardado con éxito el libro: {output_path}")
    except PermissionError:
        print(f"\n[ERROR CRÍTICO] Microsoft Excel tiene abierto el archivo '{book_name}.xlsx'.")
        print("POR FAVOR CIERRE EL EJECUTABLE DE EXCEL O EL ARCHIVO E INICIE LA BÚSQUEDA NUEVAMENTE.")
        raise RuntimeError(f"El archivo {book_name}.xlsx está abierto o bloqueado. Ciérrelo primero.")


# ─────────────────────────────────────────────
# BLOQUE PRINCIPAL
# ─────────────────────────────────────────────

def run_scraper():
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

    print("=" * 60)
    print("  AUTOMATIZACIÓN SANTAMARÍA — Generador de Flowcards")
    print(f"  Fecha: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}")
    print("=" * 60)

    # Cálculo de total de tareas para el porcentaje
    total_tasks = sum(len(config["feeds"]) for config in SITES.values())
    total_tasks += len(PORTAFOLIO_CONFIG.get("feeds_grouped", {}))
    current_task = 0

    # Directorio actual
    work_dir = os.path.dirname(os.path.abspath(__file__))
    if not os.path.exists(work_dir):
        work_dir = "./"

    history_file = os.path.join(work_dir, "historial_urls.json")
    global_seen_urls = set()
    if os.path.exists(history_file):
        try:
            with open(history_file, "r", encoding="utf-8") as f:
                global_seen_urls = set(json.load(f))
        except Exception:
            pass

    # ── El Tiempo (feeds simples, con filtro de suscriptores) ──
    for site_name, config in SITES.items():
        print(f"\n\n--- INICIANDO EXTRACCIÓN PARA: {site_name} ---")
        out_path = os.path.join(work_dir, config["output_file"])
        url_base = config["url_base"]
        is_eltiempo = ("eltiempo" in url_base)
        
        articles = []
        # Conjunto de deduplicacion propio para El Tiempo (copia del historial global)
        seen_urls_et = set(global_seen_urls)
        seen_titles_et = set()  # Dedup por titulo para El Tiempo
        for cat_label, url in config["feeds"].items():
            pct = int((current_task / total_tasks) * 100)
            print(f"\nEspere un momento, estamos buscando noticias para sus Flowcards. ({pct}%)")
            art = fetch_top_article(url, cat_label, url_base, seen_urls_et, check_paywall=is_eltiempo)
            current_task += 1
            if art:
                title_key = art["titulo_raw"].strip().lower()
                if title_key not in seen_titles_et:
                    seen_titles_et.add(title_key)
                    articles.append(art)
                else:
                    print(f"   [!] Omitido (titulo duplicado entre categorias): {art['titulo_raw'][:60]}")
        # Agregar las URLs nuevas de El Tiempo al historial global
        global_seen_urls.update(seen_urls_et)
                
        if not articles:
            print(f"\n[!] No se encontraron articulos aptos para {site_name}.")
            continue
            
        export_to_excel(articles, out_path, site_name)

    # ── Portafolio (feeds agrupados, con filtro de suscriptores) ──
    print(f"\n\n--- INICIANDO EXTRACCIÓN PARA: Portafolio ---")
    p_config = PORTAFOLIO_CONFIG
    p_out_path = os.path.join(work_dir, str(p_config["output_file"]))
    p_url_base = str(p_config["url_base"])
    is_portafolio = ("portafolio" in p_url_base.lower())
    
    p_articles = []
    # Conjunto de deduplicacion PROPIO para Portafolio (independiente de El Tiempo)
    p_seen_urls = set(global_seen_urls)
    p_seen_titles = set()  # Dedup por titulo entre grupos de Portafolio
    feeds_grouped_dict: dict = p_config.get("feeds_grouped", {})  # type: ignore
    for group_label, feed_list in feeds_grouped_dict.items():
        pct = int((current_task / total_tasks) * 100)
        print(f"\nEspere un momento, estamos buscando noticias para sus Flowcards. ({pct}%)")
        art = fetch_top_from_group(feed_list, group_label, p_url_base, p_seen_urls, check_paywall=is_portafolio)
        current_task += 1
        if art:
            title_key = art["titulo_raw"].strip().lower()
            if title_key not in p_seen_titles:
                p_seen_titles.add(title_key)
                p_articles.append(art)
            else:
                print(f"   [!] Omitido (titulo duplicado entre grupos Portafolio): {art['titulo_raw'][:60]}")
    # Agregar las URLs nuevas de Portafolio al historial global
    global_seen_urls.update(p_seen_urls)
    
    # 100% al finalizar las búsquedas
    print(f"\nEspere un momento, estamos buscando noticias para sus Flowcards. (100%)")

    if p_articles:
        export_to_excel(p_articles, p_out_path, "Portafolio")
    else:
        print(f"\n[!] No se encontraron articulos aptos para Portafolio.")

    try:
        with open(history_file, "w", encoding="utf-8") as f:
            json.dump(list(global_seen_urls), f, indent=4)
    except Exception as e:
        print(f"\n[!] No se pudo guardar el historial: {e}")

def run_scraper_selected(selected_sources=None, process_type: str = "both", include_amp: bool = True):
    if not selected_sources:
        selected_sources = ["El Tiempo", "Portafolio"]

    run_flowcards = process_type in ("flowcards", "both")
    run_temas = process_type in ("temas_del_dia", "both")

    run_el_tiempo = "El Tiempo" in selected_sources
    run_portafolio = "Portafolio" in selected_sources

    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

    print("=" * 60)
    print("  FLOWCATS v2.0 - Generador de Noticias & Radar Editorial")
    print(f"  Fecha: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}")
    print(f"  Modo de Proceso: {process_type.upper()} | Fuentes: {', '.join(selected_sources)}")
    print("=" * 60)

    total_tasks = 0
    if run_el_tiempo:
        total_tasks += sum(len(config["feeds"]) for config in SITES.values())
    if run_portafolio:
        total_tasks += len(PORTAFOLIO_CONFIG.get("feeds_grouped", {}))
    if total_tasks == 0:
        print(f"\n[!] No se seleccionó ningún medio para procesar.")
        return {"El Tiempo": [], "Portafolio": [], "top": []}

    current_task = 0
    work_dir = os.path.dirname(os.path.abspath(__file__))
    if not os.path.exists(work_dir):
        work_dir = "./"

    history_file = os.path.join(work_dir, "historial_urls.json")
    global_seen_urls = set()
    if os.path.exists(history_file):
        try:
            with open(history_file, "r", encoding="utf-8") as f:
                global_seen_urls = set(json.load(f))
        except Exception:
            pass

    articles = []
    if run_el_tiempo:
        for site_name, config in SITES.items():
            print(f"\n\n--- INICIANDO EXTRACCIÓN PARA: {site_name} ---")
            out_path = os.path.join(work_dir, config["output_file"])
            url_base = config["url_base"]
            is_eltiempo = ("eltiempo" in url_base)

            seen_urls_et = set(global_seen_urls)
            seen_titles_et = set()
            for cat_label, url in config["feeds"].items():
                pct = int((current_task / total_tasks) * 100)
                print(f"\nEspere un momento, buscando noticias para {site_name} - {cat_label} ({pct}%)")
                art = fetch_top_article(url, cat_label, url_base, seen_urls_et, check_paywall=is_eltiempo)
                current_task += 1
                if art:
                    title_key = art["titulo_raw"].strip().lower()
                    if title_key not in seen_titles_et:
                        seen_titles_et.add(title_key)
                        articles.append(art)
                    else:
                        print(f"   [!] Omitido (duplicado): {art['titulo_raw'][:60]}")
            global_seen_urls.update(seen_urls_et)

            if articles:
                if run_flowcards:
                    export_to_excel(articles, out_path, site_name, include_amp=include_amp)
                    print(f"\n[+] Archivo {site_name}.xlsx generado exitosamente con {len(articles)} Flowcards.")
            else:
                print(f"\n[!] No se encontró artículos aptos para {site_name}.")

    p_articles = []
    if run_portafolio:
        print(f"\n\n--- INICIANDO EXTRACCIÓN PARA: Portafolio ---")
        p_config = PORTAFOLIO_CONFIG
        p_out_path = os.path.join(work_dir, str(p_config["output_file"]))
        p_url_base = str(p_config["url_base"])
        is_portafolio = ("portafolio" in p_url_base.lower())

        p_seen_urls = set(global_seen_urls)
        p_seen_titles = set()
        feeds_grouped_dict: dict = p_config.get("feeds_grouped", {})
        for group_label, feed_list in feeds_grouped_dict.items():
            pct = int((current_task / total_tasks) * 100)
            print(f"\nEspere un momento, buscando noticias para Portafolio - {group_label} ({pct}%)")
            art = fetch_top_from_group(feed_list, group_label, p_url_base, p_seen_urls, check_paywall=is_portafolio)
            current_task += 1
            if art:
                title_key = art["titulo_raw"].strip().lower()
                if title_key not in p_seen_titles:
                    p_seen_titles.add(title_key)
                    p_articles.append(art)
                else:
                    print(f"   [!] Omitido (duplicado Portafolio): {art['titulo_raw'][:60]}")
        global_seen_urls.update(p_seen_urls)

        if p_articles:
            if run_flowcards:
                export_to_excel(p_articles, p_out_path, "Portafolio", include_amp=include_amp)
                print(f"\n[+] Archivo Portafolio.xlsx generado exitosamente con {len(p_articles)} Flowcards.")
        else:
            print(f"\n[!] No se encontraron artículos aptos para Portafolio.")

    print(f"\nEspere un momento, finalizando proceso. (100%)")

    try:
        with open(history_file, "w", encoding="utf-8") as f:
            json.dump(list(global_seen_urls), f, indent=4)
    except Exception as e:
        print(f"\n[!] No se pudo guardar el historial: {e}")

    # Bug 5 fix: export_temas_del_dia se ejecuta INDEPENDIENTEMENTE del medio seleccionado.
    # Toma los artículos de El Tiempo si existen, o los de Portafolio como fallback.
    top_8 = []
    if run_temas:
        all_articles_for_temas = articles if articles else p_articles
        if all_articles_for_temas:
            export_temas_del_dia(all_articles_for_temas, work_dir)
            top_8 = sorted(
                all_articles_for_temas,
                key=lambda x: (x.get("seo_score", 0), x.get("timestamp", 0)),
                reverse=True
            )[:8]
        else:
            print("\n[!] No hay artículos disponibles para generar Temas del Día.")

    return {
        "El Tiempo": articles if run_el_tiempo else [],
        "Portafolio": p_articles if run_portafolio else [],
        "top": top_8
    }

import threading
import os

try:
    import customtkinter as ctk
    from PIL import Image
    HAS_GUI = True
except (ImportError, ModuleNotFoundError):
    ctk = None
    Image = None
    HAS_GUI = False

class RealtimeLog:
    def __init__(self, tbox, progress_callback):
        self.tbox = tbox
        self.progress_callback = progress_callback
        
    def write(self, message):
        if message:
            self.tbox.after(0, self._insert_text, str(message))
            import re
            match = re.search(r'\(([0-9]+)%\)', str(message))
            if match:
                try:
                    pct = int(match.group(1))
                    if self.progress_callback:
                        self.tbox.after(0, self.progress_callback, pct)
                except Exception:
                    pass

    def _insert_text(self, message):
        self.tbox.configure(state="normal")
        self.tbox.insert("end", message)
        self.tbox.see("end")
        self.tbox.configure(state="disabled")

    def flush(self):
        pass

_BaseAppClass = ctk.CTk if ctk is not None else object

class FlowcatsApp(_BaseAppClass):
    """
    Interfaz Gráfica de Escritorio Flowcats v2.0
    Diseño editorial Dark Slate idéntico a la versión Web.
    """
    def __init__(self):
        if not HAS_GUI or ctk is None:
            raise RuntimeError("CustomTkinter / Tkinter no está disponible.")
        super().__init__()
        
        # Paleta de colores editorial Web
        self.C_BG0 = "#0B0F19"
        self.C_BG1 = "#111827"
        self.C_PANEL = "#161F30"
        self.C_BORDER = "#22304A"
        self.C_SPRING = "#00FA9A"
        self.C_EMERALD = "#10B981"
        self.C_PAPER = "#F2EEE3"
        self.C_MUTED = "#94A3B8"
        self.C_FAINT = "#64748B"
        self.C_AMBER = "#FBBF24"
        self.C_BLUE = "#3B82F6"
        self.C_DARK_INK = "#070B14"

        # Configuración de ventana principal
        self.title("Flowcats v2.0 — Sala de Redacción Automatizada")
        self.geometry("980x760")
        self.minsize(880, 680)
        self.configure(fg_color=self.C_BG0)
        ctk.set_appearance_mode("dark")
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # ── 1. CABECERA EDITORIAL (MASTHEAD) ──
        self.header_frame = ctk.CTkFrame(
            self,
            fg_color=self.C_BG1,
            corner_radius=10,
            border_width=1,
            border_color=self.C_BORDER
        )
        self.header_frame.grid(row=0, column=0, padx=18, pady=(16, 10), sticky="ew")
        self.header_frame.grid_columnconfigure(1, weight=1)

        # Logo / Stamp
        try:
            if getattr(sys, 'frozen', False):
                base_path = sys._MEIPASS # type: ignore
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
            
            logo_path = os.path.join(base_path, "logo_flowcats.png")
            if os.path.exists(logo_path):
                logo_img = ctk.CTkImage(
                    light_image=Image.open(logo_path),
                    dark_image=Image.open(logo_path),
                    size=(56, 56)
                )
                self.logo_label = ctk.CTkLabel(self.header_frame, image=logo_img, text="")
            else:
                self.logo_label = ctk.CTkLabel(self.header_frame, text="🐈‍⬛", font=("Segoe UI Emoji", 38))
        except Exception:
            self.logo_label = ctk.CTkLabel(self.header_frame, text="🐈‍⬛", font=("Segoe UI Emoji", 38))
        self.logo_label.grid(row=0, column=0, rowspan=2, padx=(16, 12), pady=12)

        # Título y Subtítulo
        title_box = ctk.CTkFrame(self.header_frame, fg_color="transparent")
        title_box.grid(row=0, column=1, rowspan=2, sticky="w", pady=10)

        title_row = ctk.CTkFrame(title_box, fg_color="transparent")
        title_row.pack(anchor="w")

        self.title_label = ctk.CTkLabel(
            title_row,
            text="FLOWCATS",
            font=("Century Gothic", 26, "bold"),
            text_color=self.C_PAPER
        )
        self.title_label.pack(side="left")

        self.ver_chip = ctk.CTkLabel(
            title_row,
            text="v2.0 DESKTOP",
            font=("Consolas", 10, "bold"),
            text_color=self.C_SPRING,
            fg_color="#064E3B",
            corner_radius=4,
            padx=7,
            pady=2
        )
        self.ver_chip.pack(side="left", padx=(10, 0))

        self.subtitle_label = ctk.CTkLabel(
            title_box,
            text="SALA DE REDACCIÓN AUTOMATIZADA · NOTICIAS & FLOWCARDS SEO",
            font=("Consolas", 9, "bold"),
            text_color=self.C_FAINT
        )
        self.subtitle_label.pack(anchor="w", pady=(3, 0))

        # Badge de Estado IA
        self.ai_badge = ctk.CTkLabel(
            self.header_frame,
            text="⚡ Groq AI (gpt-oss-120b) Activo",
            font=("Consolas", 11, "bold"),
            text_color="#6EE7B7",
            fg_color="#064E3B",
            corner_radius=6,
            padx=12,
            pady=6
        )
        self.ai_badge.grid(row=0, column=2, rowspan=2, padx=16, pady=12, sticky="e")

        # ── 2. CUBIERTA / PANELES DE CONTROL (2 COLUMNAS) ──
        self.deck_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.deck_frame.grid(row=1, column=0, padx=18, pady=0, sticky="ew")
        self.deck_frame.grid_columnconfigure(0, weight=5)
        self.deck_frame.grid_columnconfigure(1, weight=4)

        # ── PANEL IZQUIERDO: CONFIGURACIÓN ──
        self.cfg_panel = ctk.CTkFrame(
            self.deck_frame,
            fg_color=self.C_BG1,
            corner_radius=10,
            border_width=1,
            border_color=self.C_BORDER
        )
        self.cfg_panel.grid(row=0, column=0, padx=(0, 8), pady=6, sticky="nsew")

        cfg_tag = ctk.CTkLabel(
            self.cfg_panel,
            text="01 · CONFIGURACIÓN DE EXTRACCIÓN",
            font=("Consolas", 10, "bold"),
            text_color=self.C_SPRING
        )
        cfg_tag.pack(anchor="w", padx=16, pady=(12, 6))

        # Tipos de proceso (Flowcards y Temas del Día)
        mode_label = ctk.CTkLabel(
            self.cfg_panel,
            text="TIPO DE PROCESO",
            font=("Consolas", 9, "bold"),
            text_color=self.C_FAINT
        )
        mode_label.pack(anchor="w", padx=16, pady=(2, 4))

        modes_box = ctk.CTkFrame(self.cfg_panel, fg_color="transparent")
        modes_box.pack(fill="x", padx=16, pady=(0, 8))

        self.flowcards_var = ctk.BooleanVar(value=True)
        self.temas_var = ctk.BooleanVar(value=True)

        self.flowcards_check = ctk.CTkCheckBox(
            modes_box,
            text="Flowcards SEO (máx 5 palabras)",
            variable=self.flowcards_var,
            font=("Century Gothic", 12, "bold"),
            fg_color=self.C_EMERALD,
            hover_color="#059669",
            text_color=self.C_PAPER,
            command=self.on_config_change
        )
        self.flowcards_check.pack(side="left", padx=(0, 16))

        self.temas_check = ctk.CTkCheckBox(
            modes_box,
            text="Temas del Día (≤ 25c)",
            variable=self.temas_var,
            font=("Century Gothic", 12, "bold"),
            fg_color=self.C_AMBER,
            hover_color="#D97706",
            text_color=self.C_PAPER,
            command=self.on_config_change
        )
        self.temas_check.pack(side="left")

        # Medios informativos
        src_label = ctk.CTkLabel(
            self.cfg_panel,
            text="MEDIO INFORMATIVO",
            font=("Consolas", 9, "bold"),
            text_color=self.C_FAINT
        )
        src_label.pack(anchor="w", padx=16, pady=(4, 4))

        sources_box = ctk.CTkFrame(self.cfg_panel, fg_color="transparent")
        sources_box.pack(fill="x", padx=16, pady=(0, 8))

        self.el_tiempo_var = ctk.BooleanVar(value=True)
        self.portafolio_var = ctk.BooleanVar(value=True)

        self.el_tiempo_check = ctk.CTkCheckBox(
            sources_box,
            text="El Tiempo (Nacional)",
            variable=self.el_tiempo_var,
            font=("Century Gothic", 12, "bold"),
            fg_color=self.C_BLUE,
            hover_color="#2563EB",
            text_color=self.C_PAPER,
            command=self.on_config_change
        )
        self.el_tiempo_check.pack(side="left", padx=(0, 16))

        self.portafolio_check = ctk.CTkCheckBox(
            sources_box,
            text="Portafolio (Economía)",
            variable=self.portafolio_var,
            font=("Century Gothic", 12, "bold"),
            fg_color=self.C_AMBER,
            hover_color="#D97706",
            text_color=self.C_PAPER,
            command=self.on_config_change
        )
        self.portafolio_check.pack(side="left")

        # AMP Switch y Botón de Inicio
        actions_row = ctk.CTkFrame(self.cfg_panel, fg_color="transparent")
        actions_row.pack(fill="x", padx=16, pady=(4, 12))

        self.amp_var = ctk.BooleanVar(value=True)
        self.amp_switch = ctk.CTkSwitch(
            actions_row,
            text="Incluir URL AMP",
            variable=self.amp_var,
            font=("Century Gothic", 11),
            text_color=self.C_MUTED,
            progress_color=self.C_BLUE
        )
        self.amp_switch.pack(side="left")

        self.btn_run = ctk.CTkButton(
            actions_row,
            text="▶ INICIAR BÚSQUEDA",
            font=("Century Gothic", 12, "bold"),
            fg_color=self.C_EMERALD,
            hover_color=self.C_SPRING,
            text_color="#04120C",
            height=34,
            corner_radius=6,
            command=self.start_thread
        )
        self.btn_run.pack(side="right")

        # ── PANEL DERECHO: ESTADO Y ACCESO A ARCHIVOS ──
        self.files_panel = ctk.CTkFrame(
            self.deck_frame,
            fg_color=self.C_BG1,
            corner_radius=10,
            border_width=1,
            border_color=self.C_BORDER
        )
        self.files_panel.grid(row=0, column=1, padx=(8, 0), pady=6, sticky="nsew")

        status_tag = ctk.CTkLabel(
            self.files_panel,
            text="02 · ESTADO Y ARCHIVOS",
            font=("Consolas", 10, "bold"),
            text_color=self.C_SPRING
        )
        status_tag.pack(anchor="w", padx=16, pady=(12, 4))

        # Status text y porcentaje
        stat_row = ctk.CTkFrame(self.files_panel, fg_color="transparent")
        stat_row.pack(fill="x", padx=16, pady=(2, 4))

        self.status_label = ctk.CTkLabel(
            stat_row,
            text="En espera de instrucciones...",
            font=("Century Gothic", 11),
            text_color=self.C_MUTED,
            anchor="w"
        )
        self.status_label.pack(side="left", fill="x", expand=True)

        self.pct_label = ctk.CTkLabel(
            stat_row,
            text="0%",
            font=("Consolas", 14, "bold"),
            text_color=self.C_PAPER
        )
        self.pct_label.pack(side="right")

        self.progress_bar = ctk.CTkProgressBar(
            self.files_panel,
            mode="determinate",
            progress_color=self.C_SPRING,
            fg_color=self.C_DARK_INK,
            height=8,
            corner_radius=4
        )
        self.progress_bar.pack(fill="x", padx=16, pady=(0, 10))
        self.progress_bar.set(0)

        # Botones de acceso rápido a Excel
        files_btn_box = ctk.CTkFrame(self.files_panel, fg_color="transparent")
        files_btn_box.pack(fill="x", padx=16, pady=(0, 10))

        self.btn_open_et = ctk.CTkButton(
            files_btn_box,
            text="📊 El Tiempo.xlsx",
            font=("Consolas", 10, "bold"),
            fg_color="#1E293B",
            hover_color="#334155",
            text_color=self.C_BLUE,
            height=28,
            corner_radius=5,
            command=lambda: self.open_excel_file("El Tiempo.xlsx")
        )
        self.btn_open_et.pack(side="left", padx=(0, 6), expand=True, fill="x")

        self.btn_open_pf = ctk.CTkButton(
            files_btn_box,
            text="📊 Portafolio.xlsx",
            font=("Consolas", 10, "bold"),
            fg_color="#1E293B",
            hover_color="#334155",
            text_color=self.C_AMBER,
            height=28,
            corner_radius=5,
            command=lambda: self.open_excel_file("Portafolio.xlsx")
        )
        self.btn_open_pf.pack(side="left", padx=(0, 6), expand=True, fill="x")

        self.btn_open_temas = ctk.CTkButton(
            files_btn_box,
            text="⭐ TEMAS DÍA.xlsx",
            font=("Consolas", 10, "bold"),
            fg_color="#1E293B",
            hover_color="#334155",
            text_color=self.C_SPRING,
            height=28,
            corner_radius=5,
            command=lambda: self.open_excel_file("TEMAS DEL DÍA.xlsx")
        )
        self.btn_open_temas.pack(side="left", expand=True, fill="x")

        # ── 3. TERMINAL EDITORIAL / CONSOLA EN VIVO ──
        self.term_frame = ctk.CTkFrame(
            self,
            fg_color=self.C_BG1,
            corner_radius=10,
            border_width=1,
            border_color=self.C_BORDER
        )
        self.term_frame.grid(row=2, column=0, padx=18, pady=(8, 16), sticky="nsew")
        self.term_frame.grid_columnconfigure(0, weight=1)
        self.term_frame.grid_rowconfigure(1, weight=1)

        # Header de la consola
        term_header = ctk.CTkFrame(self.term_frame, fg_color=self.C_DARK_INK, corner_radius=6, height=32)
        term_header.grid(row=0, column=0, padx=8, pady=(8, 4), sticky="ew")

        dots_label = ctk.CTkLabel(
            term_header,
            text="🔴 🟡 🟢",
            font=("Segoe UI Emoji", 10)
        )
        dots_label.pack(side="left", padx=10)

        term_title = ctk.CTkLabel(
            term_header,
            text="MESA DE REDACCIÓN · CONSOLA EN VIVO",
            font=("Consolas", 10, "bold"),
            text_color=self.C_FAINT
        )
        term_title.pack(side="left", padx=6)

        self.btn_clear_log = ctk.CTkButton(
            term_header,
            text="Limpiar",
            font=("Consolas", 9),
            fg_color="transparent",
            hover_color="#1E293B",
            text_color=self.C_MUTED,
            width=50,
            height=20,
            command=self.clear_logs
        )
        self.btn_clear_log.pack(side="right", padx=10)

        # Cuerpo del Log
        self.log_box = ctk.CTkTextbox(
            self.term_frame,
            font=("Consolas", 11),
            fg_color=self.C_DARK_INK,
            text_color=self.C_SPRING,
            corner_radius=6
        )
        self.log_box.grid(row=1, column=0, padx=8, pady=(0, 8), sticky="nsew")
        self.log_box.configure(state="disabled")

        self.original_stdout = sys.stdout
        self.refresh_analysis_status()
        self.check_generated_files()

    def refresh_analysis_status(self):
        if get_groq_api_key():
            self.ai_badge.configure(
                text="⚡ Groq AI (gpt-oss-120b) Activo",
                fg_color="#064E3B",
                text_color="#6EE7B7"
            )
            return "Groq AI"
        else:
            self.ai_badge.configure(
                text="⚙ Modo Heurístico",
                fg_color="#78350F",
                text_color=self.C_AMBER
            )
            return "Heurístico"

    def on_config_change(self):
        has_sources = self.el_tiempo_var.get() or self.portafolio_var.get()
        has_modes = self.flowcards_var.get() or self.temas_var.get()

        if has_sources and has_modes:
            self.btn_run.configure(state="normal", fg_color=self.C_EMERALD)
            self.status_label.configure(text="Listo para ejecutar", text_color=self.C_MUTED)
        else:
            self.btn_run.configure(state="disabled", fg_color="#334155")
            if not has_sources:
                self.status_label.configure(text="Seleccione al menos un medio", text_color=self.C_AMBER)
            elif not has_modes:
                self.status_label.configure(text="Seleccione al menos un tipo de proceso", text_color=self.C_AMBER)

    def get_selected_sources(self):
        sources = []
        if self.el_tiempo_var.get():
            sources.append("El Tiempo")
        if self.portafolio_var.get():
            sources.append("Portafolio")
        return sources

    def get_process_type(self):
        fc = self.flowcards_var.get()
        td = self.temas_var.get()
        if fc and td:
            return "both"
        elif td:
            return "temas_del_dia"
        return "flowcards"

    def clear_logs(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def open_excel_file(self, filename: str):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(base_dir, filename)
        if os.path.exists(path):
            try:
                os.startfile(path)
            except Exception as e:
                print(f"[!] No se pudo abrir {filename}: {e}")
        else:
            print(f"[!] El archivo '{filename}' aún no ha sido generado.")

    def check_generated_files(self):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        
        if os.path.exists(os.path.join(base_dir, "El Tiempo.xlsx")):
            self.btn_open_et.configure(fg_color="#1E3A5F", text_color="#60A5FA")
        if os.path.exists(os.path.join(base_dir, "Portafolio.xlsx")):
            self.btn_open_pf.configure(fg_color="#5F3A1E", text_color=self.C_AMBER)
        if os.path.exists(os.path.join(base_dir, "TEMAS DEL DÍA.xlsx")) or os.path.exists(os.path.join(base_dir, "TEMAS DEL DIA.xlsx")):
            self.btn_open_temas.configure(fg_color="#064E3B", text_color=self.C_SPRING)

    def update_progress(self, pct):
        self.progress_bar.set(pct / 100.0)
        self.pct_label.configure(text=f"{pct}%")
        if pct == 100:
            self.status_label.configure(text="✔ Proceso completado exitosamente", text_color=self.C_SPRING)
            self.btn_run.configure(state="normal", text="▶ INICIAR BÚSQUEDA", fg_color=self.C_EMERALD)
            self.check_generated_files()

    def start_thread(self):
        sources = self.get_selected_sources()
        if not sources:
            return
        
        process_type = self.get_process_type()
        include_amp = self.amp_var.get()
        ai_mode = self.refresh_analysis_status()

        self.btn_run.configure(state="disabled", text="⏳ PROCESANDO...", fg_color="#334155")
        self.status_label.configure(text=f"Extrayendo noticias ({ai_mode})...", text_color=self.C_PAPER)
        self.progress_bar.set(0.0)
        self.pct_label.configure(text="0%")
        self.clear_logs()

        # Redireccionar salida estándar a la consola visual
        sys.stdout = RealtimeLog(self.log_box, self.update_progress) # type: ignore

        # Lanzar tarea en segundo plano
        self.thread = threading.Thread(
            target=self.run_process,
            args=(sources, process_type, include_amp),
            daemon=True
        )
        self.thread.start()

    def run_process(self, sources, process_type, include_amp):
        try:
            run_scraper_selected(sources, process_type=process_type, include_amp=include_amp)
        except Exception as e:
            print(f"\n[ERROR] Ocurrió un fallo en la extracción: {e}")
            self.status_label.configure(text="Error en la ejecución", text_color="#F87171")
            self.btn_run.configure(state="normal", text="Reintentar", fg_color=self.C_EMERALD)
        finally:
            sys.stdout = self.original_stdout

    def on_closing(self):
        try:
            sys.stdout = self.original_stdout
        except Exception:
            pass
        self.destroy()

if __name__ == "__main__":
    if HAS_GUI and ctk is not None:
        try:
            app = FlowcatsApp()
            app.protocol("WM_DELETE_WINDOW", app.on_closing)
            app.mainloop()
        except Exception as e:
            print(f"[!] Error al iniciar interfaz gráfica: {e}")
            print("[*] Ejecutando extracción en modo consola (CLI)...")
            run_scraper_selected(["El Tiempo", "Portafolio"], include_amp=True)
    else:
        print("[!] Interfaz gráfica no disponible en este entorno (falta _tkinter / CustomTkinter).")
        print("[*] Ejecutando extracción en modo consola (CLI)...")
        run_scraper_selected(["El Tiempo", "Portafolio"], include_amp=True)

