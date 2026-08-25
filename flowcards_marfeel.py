# -*- coding: utf-8 -*-
"""
flowcards_marfeel.py
====================
Automatiza la actualización de Flowcards SEO en hub.marfeel.com/compass/editorial
para el medio El Tiempo.

Por cada fila del Excel "El Tiempo.xlsx":
  1. Ubica el Flowcard SEO de la categoría de la nota.
  2. Actualiza el campo "Document URL" (pestaña Content).
  3. Actualiza el filtro "URL !=" (pestaña Targeting).
  4. Edita el título del widget (texto después del último "|").
  5. Verifica la URL y publica con "Save changes & publish".

Uso:
    python flowcards_marfeel.py                       # Procesa todas las filas
    python flowcards_marfeel.py --login               # Solo guarda/renueva la sesión
    python flowcards_marfeel.py --debug               # Navegador visible + pausas en error
    python flowcards_marfeel.py --categoria Salud     # Solo procesa una categoría

Dependencias:
    pip install playwright pandas openpyxl
    playwright install chromium
"""

import argparse
import json
import logging
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
except ImportError:
    pd = None  # type: ignore

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

from playwright.sync_api import (
    Browser,
    BrowserContext,
    Page,
    Playwright,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "dist" / "El Tiempo.xlsx"
SESSION_FILE = BASE_DIR / "marfeel_session.json"
LOG_FILE = BASE_DIR / "flowcards_log.txt"

MARFEEL_URL = "https://hub.marfeel.com/compass/editorial"
MARFEEL_LOGIN_URL = "https://hub.marfeel.com"

# Timeouts (milisegundos)
TIMEOUT_NAVIGATION = 30_000
TIMEOUT_ELEMENT = 15_000
TIMEOUT_SEARCH_RESULTS = 10_000

# Pausa entre acciones para evitar rate-limiting (segundos)
INTER_ACTION_DELAY = 0.8

# Mapeo de nombre de categoría del Excel → fragmento del nombre del Flowcard en Marfeel
# El script buscará un Flowcard cuyo título contenga "Flowcard SEO <categoria>"
# ignorando mayúsculas/minúsculas.
CATEGORY_TO_FLOWCARD: dict[str, str] = {
    "Bogota":   "Flowcard SEO Bogota",
    "Colombia": "Flowcard SEO Colombia",
    "Mundo":    "Flowcard SEO Mundo",
    "Economia": "Flowcard SEO Economia",
    "Deportes": "Flowcard SEO Deportes",
    "Politica": "Flowcard SEO Politica",
    "Cultura":  "Flowcard SEO Cultura",
    "Vida":     "Flowcard SEO Vida",
    "Justicia": "Flowcard SEO Justicia",
    "Salud":    "Flowcard SEO Salud",
}

# ─────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────

def setup_logging(debug: bool = False) -> logging.Logger:
    """Configura logging a consola y a archivo."""
    logger = logging.getLogger("flowcards")
    logger.setLevel(logging.DEBUG if debug else logging.INFO)

    fmt = logging.Formatter(
        "[%(asctime)s] %(levelname)s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # Consola
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    # Archivo
    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return logger


logger = logging.getLogger("flowcards")

# ─────────────────────────────────────────────
# LECTURA DEL EXCEL
# ─────────────────────────────────────────────

def load_excel_data(categoria_filter: str | None = None) -> list[dict]:
    """
    Lee El Tiempo.xlsx y devuelve una lista de dicts con:
      { "url": str, "titulo": str, "categoria": str }

    Si categoria_filter está definido, solo devuelve las filas de esa categoría.
    Se procesan TODAS las filas (con o sin ✔).
    """
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"No se encontró el Excel en: {EXCEL_PATH}")

    if pd is not None:
        df = pd.read_excel(EXCEL_PATH, engine="openpyxl")
        # Normalizar nombres de columnas (quitar espacios extras)
        df.columns = [str(c).strip() for c in df.columns]

        # Detectar columnas por nombre (flexible)
        # "URL SIN AMP" es la columna de URLs sin AMP; priorizarla sobre otras con "URL"
        url_col = next((c for c in df.columns if "URL SIN" in c.upper()), None)
        if not url_col:
            url_col = next((c for c in df.columns if "URL" in c.upper() and "URL AMP" not in c.upper()), None)
        titulo_col = next(
            (c for c in df.columns if c.upper().replace("Í", "I").replace("Ú", "U").startswith("TITULO")
             or c.upper().startswith("TITULO") or c.upper().startswith("TÍTULO")),
            None
        )
        if not titulo_col:
            titulo_col = next((c for c in df.columns if "TITULO" in c.upper() or "TÍTULO" in c.upper()), None)
        cat_col = next((c for c in df.columns if "CATEG" in c.upper()), None)

        if not url_col or not titulo_col or not cat_col:
            raise ValueError(
                f"No se encontraron las columnas necesarias. Encontradas: {list(df.columns)}"
            )

        rows = []
        for _, row in df.iterrows():
            url = str(row[url_col]).strip() if pd.notna(row[url_col]) else ""
            titulo = str(row[titulo_col]).strip() if pd.notna(row[titulo_col]) else ""
            categoria = str(row[cat_col]).strip() if pd.notna(row[cat_col]) else ""

            # Saltar filas vacías o encabezado repetido
            if not url or url.lower() == "url sin amp" or not url.startswith("http"):
                continue
            if not categoria or categoria.lower() == "categoría":
                continue

            if categoria_filter and categoria.lower() != categoria_filter.lower():
                continue

            rows.append({"url": url, "titulo": titulo, "categoria": categoria})

        return rows

    # Fallback con openpyxl si pandas no está disponible
    if openpyxl is None:
        raise ImportError("Se requiere pandas o openpyxl. Instala con: pip install pandas openpyxl")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [str(ws.cell(1, c).value or "").strip().upper() for c in range(1, ws.max_column + 1)]

    def find_col(keyword: str) -> int | None:
        for i, h in enumerate(headers):
            if keyword in h:
                return i + 1
        return None

    url_col = find_col("URL SIN") or find_col("URL")
    titulo_col = find_col("TITULO") or find_col("TÍTULO")
    cat_col = find_col("CATEG")

    if not url_col or not titulo_col or not cat_col:
        raise ValueError(f"Columnas no encontradas. Encabezados: {headers}")

    rows = []
    for r in range(2, ws.max_row + 1):
        url = str(ws.cell(r, url_col).value or "").strip()
        titulo = str(ws.cell(r, titulo_col).value or "").strip()
        categoria = str(ws.cell(r, cat_col).value or "").strip()

        if not url or not url.startswith("http"):
            continue
        if categoria_filter and categoria.lower() != categoria_filter.lower():
            continue

        rows.append({"url": url, "titulo": titulo, "categoria": categoria})

    return rows


# ─────────────────────────────────────────────
# GESTIÓN DE SESIÓN
# ─────────────────────────────────────────────

def do_manual_login(playwright: Playwright) -> None:
    """
    Abre un navegador visible para que el usuario haga login manualmente.
    Guarda el estado de la sesión en SESSION_FILE.
    """
    logger.info("Iniciando sesión manual. Se abrirá el navegador...")
    logger.info("Por favor, inicia sesión en Marfeel. El script esperará hasta que estés en el dashboard.")

    browser = playwright.chromium.launch(headless=False, slow_mo=50)
    context = browser.new_context()
    page = context.new_page()

    page.goto(MARFEEL_LOGIN_URL, timeout=TIMEOUT_NAVIGATION)

    logger.info("Esperando que completes el login... (tienes 3 minutos)")
    logger.info("Una vez en el dashboard de Marfeel, REGRESA A ESTA CONSOLA y presiona ENTER.")
    input("  >> Presiona ENTER cuando hayas completado el login: ")

    # Verificar que estamos autenticados (debe haber algún elemento del dashboard)
    current_url = page.url
    logger.info(f"URL actual: {current_url}")

    # Guardar sesión
    state = context.storage_state()
    with open(SESSION_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)

    logger.info(f"Sesión guardada en: {SESSION_FILE}")
    browser.close()


def create_context(playwright: Playwright, headless: bool = True) -> tuple[Browser, BrowserContext]:
    """
    Crea un contexto de navegador.
    Si existe SESSION_FILE, lo usa para restaurar la sesión autenticada.
    """
    browser = playwright.chromium.launch(
        headless=headless,
        slow_mo=100 if not headless else 50,
        args=["--disable-blink-features=AutomationControlled"],
    )

    if SESSION_FILE.exists():
        logger.info(f"Cargando sesión desde: {SESSION_FILE}")
        context = browser.new_context(
            storage_state=str(SESSION_FILE),
            viewport={"width": 1600, "height": 900},
        )
    else:
        logger.warning("No se encontró sesión guardada. Ejecuta con --login primero.")
        context = browser.new_context(viewport={"width": 1600, "height": 900})

    return browser, context


# ─────────────────────────────────────────────
# NAVEGACIÓN A FLOWCARDS
# ─────────────────────────────────────────────

def navigate_to_editorial(page: Page) -> None:
    """Navega a la sección Editorial de Marfeel."""
    logger.debug("Navegando a editorial de Marfeel...")
    page.goto(MARFEEL_URL, timeout=TIMEOUT_NAVIGATION)

    # Esperar a que la página cargue (buscar algún elemento del sidebar)
    try:
        page.wait_for_load_state("networkidle", timeout=TIMEOUT_NAVIGATION)
    except PlaywrightTimeoutError:
        logger.debug("Timeout en networkidle, continuando...")


def click_flowcards_icon(page: Page) -> None:
    """
    Hace click en el ícono de rayo (Flowcards), que es el tercer ícono
    en el menú lateral izquierdo de la sección Editorial.

    Estrategia: Buscar por aria-label, title, o texto asociado.
    """
    logger.debug("Buscando ícono de Flowcards (rayo)...")

    # Intentar múltiples selectores en orden de preferencia
    selectors = [
        # Por aria-label o title que mencione flowcard/lightning
        '[aria-label*="lowcard" i]',
        '[title*="lowcard" i]',
        '[aria-label*="lightning" i]',
        # Por href que contenga "flowcard"
        'a[href*="flowcard" i]',
        # SVG de rayo típico (Heroicons / FontAwesome lightning bolt)
        'svg[data-icon="bolt"]',
        # Tercer ítem del menú lateral
        'nav li:nth-child(3) a',
        'aside li:nth-child(3) a',
        # Fallback: cualquier elemento que contenga "Flowcard" en su texto
        'text=Flowcard',
    ]

    clicked = False
    for sel in selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=2000):
                el.click()
                logger.debug(f"Click en ícono Flowcards con selector: {sel}")
                clicked = True
                break
        except Exception:
            continue

    if not clicked:
        # Último recurso: buscar por posición (tercer ícono del nav lateral)
        logger.warning("No se encontró el ícono de rayo por selector. Intentando por posición...")
        try:
            # Buscar íconos del sidebar y hacer click en el tercero
            icons = page.locator("nav a, aside a, [role='navigation'] a").all()
            if len(icons) >= 3:
                icons[2].click()
                clicked = True
            else:
                raise ValueError("No hay suficientes íconos en el sidebar")
        except Exception as e:
            raise RuntimeError(
                f"No se pudo hacer click en el ícono de Flowcards. "
                f"Ejecuta con --debug para inspeccionar la página. Error: {e}"
            )

    time.sleep(INTER_ACTION_DELAY)
    page.wait_for_load_state("domcontentloaded", timeout=TIMEOUT_ELEMENT)


def find_and_open_flowcard(page: Page, categoria: str) -> bool:
    """
    Busca el Flowcard SEO de la categoría dada.
    Ignora los que dicen "Portafolio".
    Hace click en él para abrirlo.

    Retorna True si lo encontró y abrió, False si no.
    """
    flowcard_name = CATEGORY_TO_FLOWCARD.get(categoria)
    if not flowcard_name:
        # Intentar construir el nombre automáticamente
        flowcard_name = f"Flowcard SEO {categoria}"
        logger.warning(f"Categoría '{categoria}' no está en el mapeo. Buscando: '{flowcard_name}'")

    logger.info(f"Buscando Flowcard: '{flowcard_name}'")

    # Esperar a que los Flowcards carguen
    try:
        page.wait_for_selector(
            "text=Flowcard SEO",
            timeout=TIMEOUT_ELEMENT,
        )
    except PlaywrightTimeoutError:
        logger.error("No se cargaron los Flowcards en la página.")
        return False

    time.sleep(INTER_ACTION_DELAY)

    # Buscar todos los elementos que contengan "Flowcard SEO"
    # y filtrar los que dicen "Portafolio"
    all_flowcards = page.locator("[class*='card'], [class*='item'], li, tr").all()

    # Estrategia principal: buscar por texto exacto del nombre del flowcard
    target_text = flowcard_name.lower()

    # Método 1: Buscar directamente el elemento con ese texto
    candidates = page.get_by_text(flowcard_name, exact=False).all()

    # Filtrar los que mencionen "Portafolio"
    for el in candidates:
        try:
            # Obtener el contenedor padre para ver el texto completo
            parent_text = ""
            try:
                # Intentar obtener el texto del contenedor padre (card completa)
                parent = el.locator("xpath=..").first
                parent_text = (parent.inner_text() or "").lower()
            except Exception:
                parent_text = (el.inner_text() or "").lower()

            if "portafolio" in parent_text.lower():
                logger.debug(f"Ignorando Flowcard con 'Portafolio': {parent_text[:80]}")
                continue

            # Verificar que el nombre coincide con nuestra categoría
            el_text = (el.inner_text() or "").strip()
            if target_text in el_text.lower() or el_text.lower() in target_text:
                logger.debug(f"Flowcard encontrado: '{el_text}'")
                el.click()
                time.sleep(INTER_ACTION_DELAY)
                page.wait_for_load_state("domcontentloaded", timeout=TIMEOUT_ELEMENT)
                return True

        except PlaywrightTimeoutError:
            continue
        except Exception as e:
            logger.debug(f"Error evaluando elemento: {e}")
            continue

    # Método 2: Buscar en toda la página con contains insensible a mayúsculas
    logger.debug("Método 1 falló, intentando método 2 (XPath)...")
    try:
        # XPath que busca el nombre exacto del flowcard (case insensitive)
        xpath_expr = f"//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚÑ', 'abcdefghijklmnopqrstuvwxyzáéíóúñ'), '{flowcard_name.lower()}') and not(contains(., 'Portafolio'))]"
        el = page.locator(f"xpath={xpath_expr}").first
        if el.is_visible(timeout=3000):
            el.click()
            time.sleep(INTER_ACTION_DELAY)
            page.wait_for_load_state("domcontentloaded", timeout=TIMEOUT_ELEMENT)
            return True
    except Exception as e:
        logger.debug(f"Método 2 falló: {e}")

    logger.error(f"No se encontró el Flowcard '{flowcard_name}' en la página.")
    return False


# ─────────────────────────────────────────────
# ACTUALIZACIÓN PESTAÑA CONTENT
# ─────────────────────────────────────────────

def update_content_tab(page: Page, new_url: str) -> None:
    """
    Navega a la pestaña 'Content' del Flowcard,
    borra el valor actual de 'Document URL' y pega la nueva URL.
    """
    logger.debug("Abriendo pestaña Content...")

    # Hacer click en la pestaña "Content"
    content_tab_selectors = [
        'button:has-text("Content")',
        '[role="tab"]:has-text("Content")',
        'a:has-text("Content")',
        'li:has-text("Content")',
        'text=Content',
    ]

    tab_clicked = False
    for sel in content_tab_selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=2000):
                el.click()
                tab_clicked = True
                logger.debug(f"Pestaña Content clickeada con: {sel}")
                break
        except Exception:
            continue

    if not tab_clicked:
        logger.warning("No se encontró la pestaña 'Content' — puede que ya esté activa.")

    time.sleep(INTER_ACTION_DELAY)

    # Buscar el campo "Document URL"
    logger.debug("Buscando campo Document URL...")

    doc_url_selectors = [
        'input[name*="document" i]',
        'input[placeholder*="url" i]',
        'input[placeholder*="document" i]',
        '[label*="Document URL" i] input',
        '[aria-label*="Document URL" i]',
        # Buscar el label y luego el input siguiente
        '//label[contains(translate(text(), "ABCDEFGHIJKLMNOPQRSTUVWXYZ", "abcdefghijklmnopqrstuvwxyz"), "document url")]//following-sibling::input',
        '//label[contains(translate(text(), "ABCDEFGHIJKLMNOPQRSTUVWXYZ", "abcdefghijklmnopqrstuvwxyz"), "document url")]/..//input',
    ]

    url_input = None
    for sel in doc_url_selectors:
        try:
            prefix = "xpath=" if sel.startswith("//") else ""
            locator = page.locator(f"{prefix}{sel}").first
            if locator.is_visible(timeout=2000):
                url_input = locator
                logger.debug(f"Campo Document URL encontrado con: {sel}")
                break
        except Exception:
            continue

    if url_input is None:
        # Fallback: buscar "Document URL" como texto y luego el primer input cercano
        try:
            label = page.get_by_text("Document URL", exact=False).first
            # Buscar el input dentro del mismo contenedor
            container = label.locator("xpath=ancestor::div[.//input][1]")
            url_input = container.locator("input").first
            if url_input.is_visible(timeout=2000):
                logger.debug("Document URL input encontrado por proximidad al label")
            else:
                url_input = None
        except Exception:
            url_input = None

    if url_input is None:
        raise RuntimeError(
            "No se encontró el campo 'Document URL'. "
            "Verifica que estás en la pestaña Content del Flowcard correcto."
        )

    # Triple click para seleccionar todo el contenido actual
    url_input.click()
    time.sleep(0.3)
    url_input.press("Control+a")
    time.sleep(0.2)
    url_input.press("Delete")
    time.sleep(0.2)

    # Pegar la nueva URL
    url_input.fill(new_url)
    time.sleep(INTER_ACTION_DELAY)

    # Verificar que se pegó correctamente
    actual_value = url_input.input_value()
    if new_url not in actual_value:
        raise RuntimeError(
            f"La URL no se pegó correctamente en Document URL. "
            f"Esperada: '{new_url}', Actual: '{actual_value}'"
        )

    logger.info(f"  ✓ Document URL actualizada: {new_url[:70]}...")


# ─────────────────────────────────────────────
# ACTUALIZACIÓN PESTAÑA TARGETING
# ─────────────────────────────────────────────

def update_targeting_tab(page: Page, new_url: str) -> None:
    """
    Navega a la pestaña 'Targeting', abre el filtro 'URL !=',
    deselecciona la URL anterior, busca y selecciona la nueva URL.
    """
    logger.debug("Abriendo pestaña Targeting...")

    targeting_tab_selectors = [
        'button:has-text("Targeting")',
        '[role="tab"]:has-text("Targeting")',
        'a:has-text("Targeting")',
        'li:has-text("Targeting")',
        'text=Targeting',
    ]

    tab_clicked = False
    for sel in targeting_tab_selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=2000):
                el.click()
                tab_clicked = True
                logger.debug(f"Pestaña Targeting clickeada con: {sel}")
                break
        except Exception:
            continue

    if not tab_clicked:
        logger.warning("No se encontró la pestaña 'Targeting' — puede que ya esté activa.")

    time.sleep(INTER_ACTION_DELAY)

    # Buscar el filtro "URL !=" — puede aparecer como "URL!=", "URL !=", "URL is not", etc.
    logger.debug("Buscando filtro 'URL !='...")

    url_filter_selectors = [
        'text=URL !=',
        'text=URL!=',
        'text=URL is not',
        '[aria-label*="URL !=" i]',
        'button:has-text("URL")',
        # Buscar cualquier elemento clickeable que contenga "URL" y "!="
        '//*[contains(text(), "URL") and contains(text(), "!=")]',
        '//*[contains(text(), "URL") and contains(text(), "!")]',
    ]

    filter_el = None
    for sel in url_filter_selectors:
        try:
            prefix = "xpath=" if sel.startswith("//") else ""
            locator = page.locator(f"{prefix}{sel}").first
            if locator.is_visible(timeout=2000):
                filter_el = locator
                logger.debug(f"Filtro URL != encontrado con: {sel}")
                break
        except Exception:
            continue

    if filter_el is None:
        raise RuntimeError(
            "No se encontró el filtro 'URL !=' en la pestaña Targeting. "
            "Ejecuta con --debug para inspeccionar la página."
        )

    filter_el.click()
    time.sleep(INTER_ACTION_DELAY * 1.5)

    # Deseleccionar la casilla actual (URL anterior)
    logger.debug("Deseleccionando URL anterior...")
    _deselect_current_url(page)

    # Buscar y seleccionar la nueva URL
    logger.debug("Buscando nueva URL en el filtro...")
    _search_and_select_url(page, new_url)

    logger.info(f"  ✓ Targeting actualizado con URL: {new_url[:70]}...")


def _deselect_current_url(page: Page) -> None:
    """
    Deselecciona la casilla marcada en la sección "Selected" del filtro.
    """
    # Buscar checkboxes marcados en la sección "Selected"
    selected_section_selectors = [
        # Sección "Selected" y luego checkboxes dentro
        '//*[contains(text(), "Selected")]/..//input[@type="checkbox" and @checked]',
        '//*[contains(text(), "Selected")]/following-sibling::*//*[@type="checkbox"]',
        # O simplemente todos los checkboxes marcados visible
    ]

    deselected = False

    # Intentar con la sección "Selected"
    for sel in selected_section_selectors:
        try:
            checkboxes = page.locator(f"xpath={sel}").all()
            for cb in checkboxes:
                if cb.is_checked() and cb.is_visible(timeout=1000):
                    cb.uncheck()
                    deselected = True
                    logger.debug("Checkbox anterior deseleccionado.")
                    break
            if deselected:
                break
        except Exception:
            continue

    if not deselected:
        # Buscar cualquier checkbox marcado visible en el panel del filtro
        try:
            all_checked = page.locator('input[type="checkbox"]:checked').all()
            for cb in all_checked:
                if cb.is_visible(timeout=1000):
                    cb.uncheck()
                    deselected = True
                    logger.debug("Checkbox anterior deseleccionado (fallback).")
                    break
        except Exception:
            pass

    if not deselected:
        logger.warning("No se encontró un checkbox marcado para deseleccionar. Puede que no haya URL anterior.")

    time.sleep(INTER_ACTION_DELAY)


def _search_and_select_url(page: Page, new_url: str) -> None:
    """
    Hace click en 'Search', pega la nueva URL, espera los resultados,
    verifica que coincide y marca la casilla. Luego confirma.
    """
    # Buscar el botón o campo "Search"
    search_selectors = [
        'button:has-text("Search")',
        'input[placeholder*="Search" i]',
        'input[type="search"]',
        '[aria-label*="Search" i]',
        'text=Search',
    ]

    search_el = None
    for sel in search_selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=2000):
                search_el = el
                logger.debug(f"Campo Search encontrado con: {sel}")
                break
        except Exception:
            continue

    if search_el is None:
        raise RuntimeError("No se encontró el campo/botón 'Search' en el filtro Targeting.")

    # Si es un botón, clickear para activar el campo de búsqueda
    tag = search_el.evaluate("el => el.tagName.toLowerCase()")
    if tag == "button":
        search_el.click()
        time.sleep(INTER_ACTION_DELAY)
        # Ahora buscar el input de búsqueda
        try:
            search_input = page.locator('input[type="search"], input[placeholder*="Search" i], input[type="text"]').first
            search_input.wait_for(state="visible", timeout=TIMEOUT_ELEMENT)
            search_el = search_input
        except Exception:
            pass

    # Pegar la URL en el campo de búsqueda
    search_el.click()
    time.sleep(0.3)
    search_el.press("Control+a")
    search_el.fill(new_url)
    time.sleep(INTER_ACTION_DELAY)

    # Esperar a que aparezcan los resultados en el dropdown
    logger.debug("Esperando resultados de búsqueda...")
    _wait_for_search_results(page, new_url)

    # Verificar y marcar el checkbox del resultado correcto
    _select_matching_result(page, new_url)

    # Click en "Confirm"
    _click_confirm(page)


def _wait_for_search_results(page: Page, expected_url: str) -> None:
    """Espera a que aparezcan resultados de búsqueda que contengan la URL esperada."""
    deadline = time.time() + TIMEOUT_SEARCH_RESULTS / 1000

    while time.time() < deadline:
        try:
            # Buscar cualquier elemento que muestre la URL en los resultados
            results = page.locator(f'text="{expected_url[:30]}"').all()
            if results:
                logger.debug(f"Resultados de búsqueda encontrados ({len(results)})")
                return

            # Alternativa: buscar un dropdown o lista de resultados
            dropdown = page.locator('[role="option"], [role="listbox"] li, [class*="dropdown"] li').all()
            if dropdown:
                return

        except Exception:
            pass

        time.sleep(0.5)

    logger.warning("Timeout esperando resultados de búsqueda. Intentando continuar...")


def _select_matching_result(page: Page, expected_url: str) -> None:
    """
    En el menú desplegable de resultados, verifica que el item coincide
    con la URL esperada y marca su checkbox.
    """
    # Buscar items del dropdown
    result_selectors = [
        '[role="option"]',
        '[role="listbox"] li',
        '[class*="dropdown"] li',
        '[class*="result"] li',
        '[class*="option"]',
    ]

    for sel in result_selectors:
        try:
            items = page.locator(sel).all()
            for item in items:
                item_text = (item.inner_text() or "").strip()
                # Verificar que la URL del resultado coincide (al menos parcialmente)
                url_slug = expected_url.split("/")[-2] if expected_url.endswith("/") else expected_url.split("/")[-1]
                if url_slug.lower() in item_text.lower() or expected_url[:50] in item_text:
                    logger.debug(f"Resultado coincidente encontrado: {item_text[:80]}")
                    # Buscar el checkbox dentro del item
                    cb = item.locator('input[type="checkbox"]').first
                    if cb.count() > 0:
                        cb.check()
                    else:
                        # Si no hay checkbox, clickear directamente el item
                        item.click()
                    time.sleep(INTER_ACTION_DELAY)
                    return

            if items:
                # Si hay items pero ninguno coincide exactamente, seleccionar el primero
                logger.warning(
                    f"Ningún resultado coincide exactamente con la URL. "
                    f"Seleccionando el primer resultado."
                )
                first = items[0]
                cb = first.locator('input[type="checkbox"]').first
                if cb.count() > 0:
                    cb.check()
                else:
                    first.click()
                time.sleep(INTER_ACTION_DELAY)
                return

        except Exception:
            continue

    # Último fallback: buscar el primer checkbox no marcado visible
    logger.warning("No se encontraron items de dropdown. Buscando cualquier checkbox...")
    try:
        checkboxes = page.locator('input[type="checkbox"]:not(:checked)').all()
        for cb in checkboxes:
            if cb.is_visible(timeout=1000):
                cb.check()
                time.sleep(INTER_ACTION_DELAY)
                return
    except Exception:
        pass

    raise RuntimeError(
        "No se pudo seleccionar la URL en el filtro Targeting. "
        "Ejecuta con --debug para inspeccionar la página."
    )


def _click_confirm(page: Page) -> None:
    """Hace click en el botón 'Confirm' del filtro Targeting."""
    confirm_selectors = [
        'button:has-text("Confirm")',
        'button:has-text("Confirmar")',
        'button:has-text("Apply")',
        '[aria-label*="Confirm" i]',
        'text=Confirm',
    ]

    for sel in confirm_selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=2000):
                el.click()
                logger.debug(f"Confirm clickeado con: {sel}")
                time.sleep(INTER_ACTION_DELAY)
                return
        except Exception:
            continue

    raise RuntimeError("No se encontró el botón 'Confirm' en el filtro Targeting.")


# ─────────────────────────────────────────────
# ACTUALIZACIÓN DEL TÍTULO DEL WIDGET
# ─────────────────────────────────────────────

def update_widget_title(page: Page, new_titulo: str) -> None:
    """
    Busca el título del Flowcard (en la parte superior izquierda),
    selecciona SOLO el texto después del último '|' y lo reemplaza con new_titulo.

    Formato esperado: "Algo | Algo | TítuloAnterior"
    Resultado:        "Algo | Algo | NuevoTitulo"
    """
    logger.debug("Buscando título del widget para editar...")

    # Buscar el elemento editable del título (usualmente un input o div contenteditable)
    title_selectors = [
        # Posiblemente sea un campo de texto en la parte superior del Flowcard
        '[contenteditable="true"]',
        'input[name*="title" i]',
        'input[name*="name" i]',
        'input[placeholder*="title" i]',
        'input[placeholder*="name" i]',
        '[aria-label*="title" i]',
        '[aria-label*="name" i]',
        # El título del widget suele aparecer en la cabecera de la card
        'h1 [contenteditable]',
        'header input',
        'header [contenteditable]',
    ]

    title_el = None
    for sel in title_selectors:
        try:
            locator = page.locator(sel).first
            if locator.is_visible(timeout=2000):
                title_el = locator
                logger.debug(f"Campo de título encontrado con: {sel}")
                break
        except Exception:
            continue

    if title_el is None:
        # Intentar hacer click directo en la parte superior izquierda de la página
        # para activar el modo edición del título
        logger.warning(
            "No se encontró el campo de título automáticamente. "
            "Intentando hacer click en la zona del título..."
        )
        try:
            # Buscar elementos que parezcan títulos de widgets (texto grande en la parte superior)
            title_candidates = page.locator("h1, h2, h3, [class*='title'], [class*='heading']").all()
            for el in title_candidates:
                text = (el.inner_text() or "").strip()
                if "|" in text and len(text) > 5:
                    el.click()
                    time.sleep(0.5)
                    # Verificar si se volvió editable
                    inner = el.locator('input, [contenteditable="true"]').first
                    if inner.count() > 0:
                        title_el = inner
                    else:
                        title_el = el
                    break
        except Exception:
            pass

    if title_el is None:
        raise RuntimeError(
            "No se pudo encontrar el campo editable del título del widget. "
            "Ejecuta con --debug para inspeccionar la página."
        )

    # Obtener el texto actual del título
    current_title = ""
    try:
        tag = title_el.evaluate("el => el.tagName.toLowerCase()")
        if tag == "input":
            current_title = title_el.input_value()
        else:
            current_title = title_el.inner_text() or ""
    except Exception:
        current_title = ""

    logger.debug(f"Título actual: '{current_title}'")

    # Determinar el nuevo título completo
    if "|" in current_title:
        # Reemplazar solo el texto después del último "|"
        parts = current_title.rsplit("|", 1)
        new_full_title = parts[0] + "| " + new_titulo
    else:
        # No hay "|", reemplazar todo
        new_full_title = new_titulo
        logger.warning(f"El título '{current_title}' no tiene '|'. Se reemplazará todo.")

    logger.debug(f"Nuevo título: '{new_full_title}'")

    # Editar el título
    try:
        tag = title_el.evaluate("el => el.tagName.toLowerCase()")
        if tag == "input":
            title_el.click()
            title_el.press("Control+a")
            title_el.fill(new_full_title)
        else:
            # contenteditable
            title_el.click()
            page.keyboard.press("Control+a")
            page.keyboard.type(new_full_title)
    except Exception as e:
        raise RuntimeError(f"Error al editar el título del widget: {e}")

    time.sleep(INTER_ACTION_DELAY)
    logger.info(f"  ✓ Título del widget actualizado: '{new_full_title}'")


# ─────────────────────────────────────────────
# VERIFICACIÓN Y PUBLICACIÓN
# ─────────────────────────────────────────────

def verify_and_publish(page: Page, expected_url: str) -> bool:
    """
    Verifica que el campo Document URL contiene la URL esperada.
    Si es correcto, hace click en 'Save changes & publish'.
    Retorna True si se publicó correctamente.
    """
    logger.debug("Verificando URL antes de publicar...")

    # Volver a la pestaña Content para verificar
    content_tab_selectors = [
        'button:has-text("Content")',
        '[role="tab"]:has-text("Content")',
        'text=Content',
    ]

    for sel in content_tab_selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=2000):
                el.click()
                time.sleep(INTER_ACTION_DELAY)
                break
        except Exception:
            continue

    # Leer el valor actual del campo Document URL
    current_url = ""
    doc_url_selectors = [
        'input[name*="document" i]',
        'input[placeholder*="url" i]',
        'input[placeholder*="document" i]',
    ]

    for sel in doc_url_selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=2000):
                current_url = el.input_value()
                break
        except Exception:
            continue

    # Verificar coincidencia
    if current_url.strip() != expected_url.strip():
        logger.error(
            f"❌ VERIFICACIÓN FALLIDA — NO se publicará.\n"
            f"   URL esperada: {expected_url}\n"
            f"   URL actual:   {current_url}"
        )
        return False

    logger.debug("Verificación OK. Publicando...")

    # Click en "Save changes & publish"
    publish_selectors = [
        'button:has-text("Save changes & publish")',
        'button:has-text("Save changes")',
        'button:has-text("Publish")',
        'button:has-text("Save")',
        '[aria-label*="publish" i]',
        '[aria-label*="save" i]',
    ]

    published = False
    for sel in publish_selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=2000):
                el.click()
                published = True
                logger.debug(f"Botón de publicar clickeado con: {sel}")
                break
        except Exception:
            continue

    if not published:
        raise RuntimeError(
            "No se encontró el botón 'Save changes & publish'. "
            "Ejecuta con --debug para inspeccionar la página."
        )

    # Esperar confirmación de guardado (toast/snackbar o cambio de estado)
    time.sleep(2)
    try:
        page.wait_for_selector(
            'text=saved, text=published, text=publicado, text=guardado, [class*="success"]',
            timeout=5000,
        )
    except Exception:
        pass  # No es crítico si el mensaje no aparece

    logger.info("  ✓ Cambios guardados y publicados.")
    return True


# ─────────────────────────────────────────────
# FUNCIÓN PRINCIPAL DE ACTUALIZACIÓN
# ─────────────────────────────────────────────

def process_row(page: Page, row: dict, debug: bool = False) -> dict:
    """
    Procesa una fila del Excel:
    Actualiza el Flowcard SEO de su categoría con la URL y título dados.

    Retorna un dict con resultado: { "status": "ok"|"error", "message": str }
    """
    url = row["url"]
    titulo = row["titulo"]
    categoria = row["categoria"]

    logger.info(f"\n{'─'*60}")
    logger.info(f"Procesando: [{categoria}] {titulo}")
    logger.info(f"  URL: {url[:80]}...")

    try:
        # 1. Navegar a editorial
        navigate_to_editorial(page)

        # 2. Click en ícono de Flowcards (rayo)
        click_flowcards_icon(page)

        # 3. Encontrar y abrir el Flowcard de la categoría
        found = find_and_open_flowcard(page, categoria)
        if not found:
            return {
                "status": "error",
                "message": f"Flowcard no encontrado para categoría '{categoria}'",
            }

        # 4. Actualizar pestaña Content (Document URL)
        update_content_tab(page, url)

        # 5. Actualizar pestaña Targeting (URL !=)
        update_targeting_tab(page, url)

        # 6. Editar título del widget
        update_widget_title(page, titulo)

        # 7. Verificar y publicar
        published = verify_and_publish(page, url)
        if not published:
            return {
                "status": "error",
                "message": "Verificación fallida: URL en el campo no coincide con la esperada. NO publicado.",
            }

        return {"status": "ok", "message": "Publicado correctamente."}

    except PlaywrightTimeoutError as e:
        msg = f"Timeout: {e}"
        logger.error(f"  ⚠ Timeout en categoría '{categoria}': {e}")
        if debug:
            logger.error("  [DEBUG] Haciendo pausa — presiona ENTER para continuar...")
            input()
        return {"status": "error", "message": msg}

    except RuntimeError as e:
        msg = str(e)
        logger.error(f"  ❌ Error en categoría '{categoria}': {msg}")
        if debug:
            logger.error("  [DEBUG] Haciendo pausa — presiona ENTER para continuar...")
            input()
        return {"status": "error", "message": msg}

    except Exception as e:
        msg = f"Error inesperado: {type(e).__name__}: {e}"
        logger.error(f"  ❌ {msg}")
        if debug:
            logger.error("  [DEBUG] Haciendo pausa — presiona ENTER para continuar...")
            input()
        return {"status": "error", "message": msg}


# ─────────────────────────────────────────────
# PUNTO DE ENTRADA
# ─────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Automatiza la actualización de Flowcards SEO en Marfeel para El Tiempo."
    )
    parser.add_argument(
        "--login",
        action="store_true",
        help="Solo guarda/renueva la sesión de Marfeel (abre navegador para login manual).",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Ejecuta con el navegador visible y hace pausas en errores.",
    )
    parser.add_argument(
        "--categoria",
        type=str,
        default=None,
        help="Procesa solo la categoría indicada (ej: --categoria Salud).",
    )
    parser.add_argument(
        "--excel",
        type=str,
        default=None,
        help="Ruta alternativa al Excel (por defecto: dist/El Tiempo.xlsx).",
    )
    args = parser.parse_args()

    # Configurar ruta del Excel si se especificó
    global EXCEL_PATH
    if args.excel:
        EXCEL_PATH = Path(args.excel)

    # Configurar logging
    global logger
    logger = setup_logging(debug=args.debug)

    logger.info("=" * 60)
    logger.info("  Automatización Flowcards SEO - El Tiempo en Marfeel")
    logger.info(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 60)

    headless = not args.debug

    with sync_playwright() as playwright:

        # ─── Modo LOGIN ───
        if args.login:
            do_manual_login(playwright)
            logger.info("✅ Sesión guardada. Ahora puedes ejecutar el script sin --login.")
            return

        # ─── Modo EJECUCIÓN ───

        # Verificar sesión
        if not SESSION_FILE.exists():
            logger.error(
                "No se encontró la sesión guardada.\n"
                "Ejecuta primero: python flowcards_marfeel.py --login"
            )
            sys.exit(1)

        # Leer Excel
        logger.info(f"Leyendo Excel: {EXCEL_PATH}")
        try:
            rows = load_excel_data(categoria_filter=args.categoria)
        except FileNotFoundError as e:
            logger.error(str(e))
            sys.exit(1)
        except Exception as e:
            logger.error(f"Error leyendo el Excel: {e}")
            sys.exit(1)

        if not rows:
            logger.warning(
                f"No se encontraron filas en el Excel"
                + (f" para la categoría '{args.categoria}'" if args.categoria else "")
                + "."
            )
            sys.exit(0)

        logger.info(f"Se procesarán {len(rows)} nota(s):")
        for r in rows:
            logger.info(f"  [{r['categoria']}] {r['titulo']}")

        # Crear navegador y ejecutar
        browser, context = create_context(playwright, headless=headless)
        page = context.new_page()
        page.set_default_timeout(TIMEOUT_ELEMENT)

        results = []
        ok_count = 0
        error_count = 0

        for i, row in enumerate(rows, 1):
            logger.info(f"\n[{i}/{len(rows)}] Procesando nota...")
            result = process_row(page, row, debug=args.debug)
            result["categoria"] = row["categoria"]
            result["url"] = row["url"]
            result["titulo"] = row["titulo"]
            results.append(result)

            if result["status"] == "ok":
                ok_count += 1
            else:
                error_count += 1

            # Pausa entre filas para no saturar Marfeel
            if i < len(rows):
                logger.info("  Esperando 3 segundos antes de la siguiente nota...")
                time.sleep(3)

        # Guardar sesión actualizada
        try:
            updated_state = context.storage_state()
            with open(SESSION_FILE, "w", encoding="utf-8") as f:
                json.dump(updated_state, f, indent=2)
            logger.debug("Sesión actualizada guardada.")
        except Exception:
            pass

        browser.close()

        # ─── RESUMEN FINAL ───
        logger.info("\n" + "=" * 60)
        logger.info("  RESUMEN FINAL")
        logger.info("=" * 60)
        logger.info(f"  ✅ Exitosos: {ok_count}")
        logger.info(f"  ❌ Errores:  {error_count}")
        logger.info(f"  Total:       {len(results)}")
        logger.info("")

        for r in results:
            status_icon = "✅" if r["status"] == "ok" else "❌"
            logger.info(f"  {status_icon} [{r['categoria']}] {r['titulo']}")
            if r["status"] == "error":
                logger.info(f"      → {r['message']}")

        logger.info(f"\nLog completo guardado en: {LOG_FILE}")

        if error_count > 0:
            sys.exit(1)


if __name__ == "__main__":
    main()
